from pathlib import Path

from openpyxl import load_workbook

CAMINHO = Path(__file__).resolve().parents[4] / "data" / "processed" / "energia_eletrica_fitted" / "KSB1 - Fitted Units 2026 - Sem Agrupamento (energia).xlsx"

FORNECEDORES = {
    "4211308770": "CEMIG DISTRIBUIÇÃO S/A",
    "4211324097": "CPFL",
    "4211333301": "COPEL DISTRIBUIÇÃO S.A",
    "4211330756": "FIAT AUTOMOVEIS S/A",
    "4211333021": "SERENA GERAÇÃO S.A (ignorar - rateio)",
}

wb = load_workbook(CAMINHO, data_only=True, read_only=True)
ws = wb.active

header = None
linhas_conta = []
linhas_fornecedor = []
contas_vistas = set()

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        header = row
        print("Cabeçalho:", header)
        continue
    conta = row[3] if len(row) > 3 else None  # Classe de custo
    fornecedor = row[5] if len(row) > 5 else None  # Fornecedor
    if conta and str(conta).strip() == "N17002S001":
        linhas_conta.append(row)
    if fornecedor and str(fornecedor).strip() in FORNECEDORES:
        linhas_fornecedor.append(row)
    if conta and "N17002" in str(conta):
        contas_vistas.add(str(conta).strip())

print(f"\nLinhas com conta N17002S001: {len(linhas_conta)}")
for r in linhas_conta[:5]:
    print(" ", r)

print(f"\nContas 'N17002*' encontradas no extrato: {sorted(contas_vistas)}")

print(f"\nLinhas com fornecedor de energia conhecido: {len(linhas_fornecedor)}")
for r in linhas_fornecedor[:10]:
    print(" ", r)
