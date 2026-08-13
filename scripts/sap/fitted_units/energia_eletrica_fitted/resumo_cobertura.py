from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

CAMINHO = Path(__file__).resolve().parents[4] / "data" / "processed" / "energia_eletrica_fitted" / "KSB1 - Fitted Units 2026 - Sem Agrupamento (energia).xlsx"

wb = load_workbook(CAMINHO, data_only=True, read_only=True)
ws = wb.active

por_centro_mes = defaultdict(list)

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    conta = row[3]
    if not conta or str(conta).strip() != "N17002S001":
        continue
    data_lanc = row[0]
    centro = row[2]
    fornecedor = row[5]
    nome = row[6]
    valor = row[16]
    mes = data_lanc.strftime("%Y-%m") if data_lanc else "?"
    por_centro_mes[centro].append((mes, fornecedor, nome, valor))

for centro in sorted(por_centro_mes):
    print(f"\n=== Centro de custo {centro} ===")
    for mes, fornecedor, nome, valor in sorted(por_centro_mes[centro]):
        print(f"  {mes}  fornecedor={fornecedor!r:15} nome={nome!r:35} valor={valor}")
