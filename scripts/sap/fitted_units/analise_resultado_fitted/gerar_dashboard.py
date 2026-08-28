"""Análise de Resultado Fitted - cockpit novo, SEPARADO do cockpit de
fechamento (pedido explícito da usuária, 2026-08-28: "não quero que as
pessoas vejam" - o cockpit de fechamento agora é usado também pela
estagiária). Gera um dashboard HTML local (sem publicar em lugar nenhum,
sem precisar de Python pra ABRIR - só pra gerar/atualizar).

v1 (escopo escolhido pela usuária): só Tendência de EBIT/Resultado.
Fonte: P&L Actual do mês mais recente já fechado, aba "Resumo Resultado
Ano" (já cobre o ano inteiro - Actual nos meses fechados, Forecast nos
meses seguintes). Linha 44 = EBIT, linha 12 = Net Sales (pra ROS%), linha
4 = meses, linha 5 = tag Actual/Forecast por mês.

Pra "atualizar sozinho mês a mês": basta rodar este script de novo depois
de cada fechamento - ele acha sozinho o P&L Actual mais recente (varre de
Dezembro pra Janeiro, pega o primeiro que existir).
"""
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "_shared"))
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "fitted_units_despesas"))
from ksb1_core import MESES_PASTA, REDE_BASE, resolver_pasta_ciclo  # noqa: E402

ANO = 2026
SAIDA_LOCAL = Path(__file__).resolve().parents[4] / "data" / "processed" / "analise_resultado_fitted"

MESES_ABREV_PT = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]


def nome_arquivo_pnl_congelado(mes: int, ano: int) -> str:
    meses_ingles = {
        1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
        7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December",
    }
    return f"{mes:02d}_P&L Fitted Units_Actual_{meses_ingles[mes]}-{ano % 100:02d}_.xlsx"


def achar_pnl_actual_mais_recente(ano: int) -> tuple[Path, int]:
    for mes in range(12, 0, -1):
        pasta_mes = REDE_BASE / str(ano) / MESES_PASTA[mes]
        pasta_ciclo = resolver_pasta_ciclo(pasta_mes, mes, "Actual")
        caminho = pasta_ciclo / nome_arquivo_pnl_congelado(mes, ano)
        if caminho.exists():
            return caminho, mes
    raise FileNotFoundError(f"Não encontrei nenhum P&L Actual congelado em {ano}.")


def extrair_resultado_ano(caminho: Path) -> dict:
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb["Resumo Resultado Ano"]

    def linha(r):
        return [ws.cell(row=r, column=c).value for c in range(4, 16)]

    tags = linha(5)
    ebit = linha(44)
    net_sales = linha(12)
    total_costs = linha(38)

    return {"tags": tags, "ebit": ebit, "net_sales": net_sales, "total_costs": total_costs}


def montar_html(dados: dict, mes_fechado: int, ano: int, caminho_fonte: Path) -> str:
    tags, ebit, net_sales = dados["tags"], dados["ebit"], dados["net_sales"]
    n_actual = sum(1 for t in tags if t == "Actual")

    ebit_ytd = sum(v for v, t in zip(ebit, tags) if t == "Actual" and isinstance(v, (int, float)))
    ebit_ultimo = ebit[n_actual - 1] if n_actual else None
    ros_valores = [
        (e / ns) for e, ns, t in zip(ebit, net_sales, tags)
        if t == "Actual" and isinstance(e, (int, float)) and isinstance(ns, (int, float)) and ns
    ]
    ros_medio = sum(ros_valores) / len(ros_valores) if ros_valores else None

    max_abs = max(abs(v) for v in ebit if isinstance(v, (int, float))) or 1
    largura_grafico = 900
    altura_grafico = 260
    n = 12
    largura_barra = (largura_grafico / n) * 0.55
    meio_y = altura_grafico / 2

    barras_svg = []
    labels_svg = []
    for i in range(n):
        v = ebit[i]
        tag = tags[i]
        cx = (largura_grafico / n) * i + (largura_grafico / n) / 2
        if isinstance(v, (int, float)):
            h = (abs(v) / max_abs) * (altura_grafico / 2 - 24)
            y = meio_y - h if v >= 0 else meio_y
            cor = "var(--cor-positivo)" if v >= 0 else "var(--cor-negativo)"
            opacidade = "1" if tag == "Actual" else "0.45"
            tracejado = "" if tag == "Actual" else 'stroke="var(--text-secondary)" stroke-width="1" stroke-dasharray="3,2"'
            barras_svg.append(
                f'<rect class="barra" data-mes="{MESES_ABREV_PT[i]}" data-valor="{v:.1f}" data-tag="{tag}" '
                f'x="{cx - largura_barra/2:.1f}" y="{y:.1f}" width="{largura_barra:.1f}" height="{h:.1f}" '
                f'rx="3" fill="{cor}" opacity="{opacidade}" {tracejado}/>'
            )
        labels_svg.append(f'<text x="{cx:.1f}" y="{altura_grafico + 16}" class="eixo-label" text-anchor="middle">{MESES_ABREV_PT[i]}</text>')

    fmt = lambda v: f"R$ {v:,.0f}K".replace(",", ".") if isinstance(v, (int, float)) else "-"
    fmt_pct = lambda v: f"{v*100:.1f}%".replace(".", ",") if isinstance(v, (int, float)) else "-"

    return f"""<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>Análise de Resultado Fitted</title>
<style>
  :root {{
    --surface-1: #fcfcfb; --page: #f2f1ee; --text-primary: #0b0b0b; --text-secondary: #52514e;
    --muted: #898781; --grid: #e1e0d9; --baseline: #c3c2b7;
    --cor-positivo: #2a78d6; --cor-negativo: #e34948;
    --border: rgba(11,11,11,0.10);
  }}
  * {{ box-sizing: border-box; }}
  body {{ margin: 0; background: var(--page); font-family: system-ui, -apple-system, "Segoe UI", sans-serif; color: var(--text-primary); }}
  .wrap {{ max-width: 1040px; margin: 0 auto; padding: 32px 24px 64px; }}
  header {{ margin-bottom: 24px; }}
  h1 {{ font-size: 22px; margin: 0 0 4px; }}
  .sub {{ color: var(--text-secondary); font-size: 13px; }}
  .kpis {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 16px; margin: 24px 0; }}
  .kpi {{ background: var(--surface-1); border: 1px solid var(--border); border-radius: 10px; padding: 18px 20px; }}
  .kpi .label {{ font-size: 12px; color: var(--text-secondary); margin-bottom: 6px; }}
  .kpi .valor {{ font-size: 28px; font-weight: 600; font-variant-numeric: tabular-nums; }}
  .kpi .valor.neg {{ color: var(--cor-negativo); }}
  .kpi .valor.pos {{ color: var(--cor-positivo); }}
  .card {{ background: var(--surface-1); border: 1px solid var(--border); border-radius: 10px; padding: 24px; }}
  .card h2 {{ font-size: 15px; margin: 0 0 4px; }}
  .legenda {{ display: flex; gap: 20px; font-size: 12px; color: var(--text-secondary); margin: 4px 0 16px; }}
  .legenda span {{ display: inline-flex; align-items: center; gap: 6px; }}
  .dot {{ width: 10px; height: 10px; border-radius: 2px; display: inline-block; }}
  svg .eixo-label {{ fill: var(--muted); font-size: 11px; }}
  svg .grid-line {{ stroke: var(--grid); stroke-width: 1; }}
  svg .baseline {{ stroke: var(--baseline); stroke-width: 1.5; }}
  .barra {{ cursor: pointer; }}
  #tooltip {{ position: fixed; display: none; background: var(--text-primary); color: #fff; font-size: 12px;
    padding: 8px 10px; border-radius: 6px; pointer-events: none; z-index: 10; white-space: nowrap; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 13px; }}
  th, td {{ text-align: right; padding: 6px 8px; border-bottom: 1px solid var(--grid); font-variant-numeric: tabular-nums; }}
  th:first-child, td:first-child {{ text-align: left; }}
  th {{ color: var(--text-secondary); font-weight: 500; }}
  .tag-forecast {{ color: var(--muted); font-size: 11px; }}
  footer {{ color: var(--muted); font-size: 11px; margin-top: 24px; }}
</style>
</head>
<body>
<div class="wrap">
  <header>
    <h1>Análise de Resultado Fitted</h1>
    <div class="sub">EBIT mensal {ano} — Actual até {MESES_ABREV_PT[mes_fechado-1]}, Forecast nos meses seguintes · fonte: {caminho_fonte.name}</div>
  </header>

  <div class="kpis">
    <div class="kpi">
      <div class="label">EBIT acumulado (Actual, Jan-{MESES_ABREV_PT[mes_fechado-1]})</div>
      <div class="valor {'pos' if ebit_ytd >= 0 else 'neg'}">{fmt(ebit_ytd)}</div>
    </div>
    <div class="kpi">
      <div class="label">EBIT último mês fechado ({MESES_ABREV_PT[mes_fechado-1]})</div>
      <div class="valor {'pos' if (ebit_ultimo or 0) >= 0 else 'neg'}">{fmt(ebit_ultimo)}</div>
    </div>
    <div class="kpi">
      <div class="label">Margem EBIT média (ROS%, Actual)</div>
      <div class="valor">{fmt_pct(ros_medio)}</div>
    </div>
  </div>

  <div class="card">
    <h2>EBIT por mês</h2>
    <div class="legenda">
      <span><span class="dot" style="background:var(--cor-positivo)"></span>Positivo</span>
      <span><span class="dot" style="background:var(--cor-negativo)"></span>Negativo</span>
      <span><span class="dot" style="background:var(--text-secondary);opacity:.45"></span>Forecast (ainda não fechado)</span>
    </div>
    <svg viewBox="0 0 {largura_grafico} {altura_grafico + 32}" width="100%" style="overflow:visible">
      <line class="baseline" x1="0" y1="{meio_y}" x2="{largura_grafico}" y2="{meio_y}"/>
      {''.join(labels_svg)}
      {''.join(barras_svg)}
    </svg>
  </div>

  <table>
    <thead><tr><th>Mês</th><th>EBIT</th><th>Net Sales</th><th>Situação</th></tr></thead>
    <tbody>
      {''.join(f"<tr><td>{MESES_ABREV_PT[i]}</td><td>{fmt(ebit[i])}</td><td>{fmt(net_sales[i])}</td><td>{'<span class=tag-forecast>Forecast</span>' if tags[i]!='Actual' else 'Actual'}</td></tr>" for i in range(12))}
    </tbody>
  </table>

  <footer>Gerado automaticamente a partir de {caminho_fonte}. Rode o script de novo depois de cada fechamento pra atualizar.</footer>
</div>
<div id="tooltip"></div>
<script>
  const tooltip = document.getElementById('tooltip');
  document.querySelectorAll('.barra').forEach(el => {{
    el.addEventListener('mousemove', (e) => {{
      const mes = el.dataset.mes, valor = parseFloat(el.dataset.valor), tag = el.dataset.tag;
      tooltip.innerHTML = `<strong>${{mes}}</strong> (${{tag}})<br>EBIT: R$ ${{valor.toLocaleString('pt-BR')}}K`;
      tooltip.style.display = 'block';
      tooltip.style.left = (e.clientX + 14) + 'px';
      tooltip.style.top = (e.clientY + 14) + 'px';
    }});
    el.addEventListener('mouseleave', () => tooltip.style.display = 'none');
  }});
</script>
</body>
</html>
"""


def gerar():
    caminho_pnl, mes_fechado = achar_pnl_actual_mais_recente(ANO)
    dados = extrair_resultado_ano(caminho_pnl)
    html = montar_html(dados, mes_fechado, ANO, caminho_pnl)

    SAIDA_LOCAL.mkdir(parents=True, exist_ok=True)
    caminho_saida = SAIDA_LOCAL / "analise_resultado_fitted.html"
    caminho_saida.write_text(html, encoding="utf-8")
    return caminho_saida, mes_fechado


if __name__ == "__main__":
    caminho, mes = gerar()
    print(f"Dashboard gerado: {caminho}")
    print(f"Mês fechado mais recente usado: {MESES_ABREV_PT[mes-1]}/{ANO}")
