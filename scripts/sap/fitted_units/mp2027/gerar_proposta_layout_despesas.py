"""MP2027 - PROPOSTA de layout pro arquivo "Detalhe_Despesas_Fitted Units"
(hoje 107 colunas, 6 blocos repetidos de "R07 JAN..DEZ" sem nome claro,
"amador" segundo a usuária). Reescreve como: Classificação (A-X, intacta)
-> bloco FORECAST (Valor Mensal / % Reajuste / Valor Final Previsão x 12
meses) -> bloco EFETIVO (Actual, x 12 meses, meses fechados preenchidos
automaticamente) -> Diferença (Forecast - Efetivo).

Mapeamento dos blocos originais do Forecast (decodificado via fórmula):
- bloco1 (col Y, 24) = Valor Mensal (base, valor estático)
- bloco4 (col BQ, 68) = % Reajuste (índice de reajuste aplicado, estático)
- bloco6 (col CQ, 94) = Valor Final Previsão (número oficial - validado
  batendo com a aba "Resumo Custos" em 2026-08-28)
Blocos 2/3/5 (quantidade, valor calculado, valor com incremento) ficam de
fora - intermediários de cálculo, não usados fora dessa aba.

EFETIVO (Actual) - lógica de atualização mês a mês (2026-08-28, ainda v1,
"depois podemos gerar uma lógica juntos por trás de toda gestorial, pra
facilitar" - usuária ciente que isso vai ser refinado):
- Fonte: arquivo oficial "KSB1 <Mês> Actual <Ano>.xlsx" (aba BASE_KSB1) do
  MÊS MAIS RECENTE JÁ FECHADO - essa aba é CUMULATIVA (acumula Jan até o
  mês do arquivo), então um único arquivo cobre todos os meses fechados
  até ali. Pra "atualizar mês a mês", basta trocar pra pasta/arquivo do
  mês mais recente quando fechar - a lógica de agregação não muda.
- Chave de casamento testada empiricamente (nenhuma bateu 100%, a melhor
  achada foi (Centro de Custo, Gestorial) = 69% de match, versus (Centro
  de Custo, Classe de Custo) = só 6% e (Centro de Custo, Nova Classe de
  Custo) = 56%. Usada (Centro de Custo, Gestorial) por ser a melhor E a
  mais estável semanticamente (é a classificação de negócio resolvida dos
  dois lados, não um código técnico de variante do SAP).
- LIMITAÇÃO CONHECIDA (comunicar sempre que usar este arquivo): como o
  Efetivo é agregado por (CC, Gestorial) e o Detalhe_Despesas tem VÁRIAS
  linhas (fornecedor/item) por combinação (CC, Gestorial), o mesmo valor
  de Efetivo aparece repetido em todas as linhas que compartilham a
  chave - somar a coluna Efetivo direto não bate com o total real
  (duplica).Útil pra comparar POR LINHA/conta, não pra somar a coluna
  inteira sem agrupar antes.
"""
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "_shared"))
from ksb1_core import nome_com_versao  # noqa: E402

ORIGEM_FORECAST = Path(r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Custos Fitted Units\Resultados Fitted\2026\07 - Jul\07_Jul_Forecast\Detalhe_Despesas_Fitted Units_Forecast July.xlsx")
ORIGEM_ACTUAL = Path(r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Custos Fitted Units\Resultados Fitted\2026\07 - Jul\07_Jul_Actual\KSB1 July Actual 2026.xlsx")
SAIDA = Path(r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Management Plan\MP2027")

MESES = ["JAN", "FEV", "MAR", "ABR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
N_MESES_FECHADOS = 7  # até Julho - muda automaticamente quando trocar ORIGEM_ACTUAL pro mês mais recente

N_COLS_CLASSIFICACAO = 24  # A-X, intactas
COLS_METADADOS = {
    "Área Responsável": 37, "Índice de Reajuste": 38,
    "Mês de Contrato": 39, "Comprador": 40, "Premissas": 41,
}
COL_VALOR_MENSAL_INICIO = 24
COL_REAJUSTE_INICIO = 68
COL_VALOR_FINAL_INICIO = 94

# paleta profissional: cinza neutro (classificação), azul petróleo (forecast),
# âmbar (efetivo/actual), verde/vermelho só na diferença (condicional)
COR_HEADER_CLASSIFICACAO = "44546A"
COR_HEADER_FORECAST = "1F4E78"
COR_HEADER_EFETIVO = "9C5700"
COR_HEADER_DIFERENCA = "375623"
COR_CEL_CLASSIFICACAO = "F2F2F2"
COR_CEL_FORECAST = "DCE6F1"
COR_CEL_EFETIVO = "FCE4D6"
COR_CEL_DIFERENCA = "E2EFDA"
FONTE_HEADER = Font(bold=True, color="FFFFFF", size=10)
FONTE_SECAO = Font(bold=True, color="FFFFFF", size=12)
BORDA_FINA = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def carregar_efetivo_por_cc_gestorial() -> dict[tuple[str, str], list[float]]:
    wb = load_workbook(ORIGEM_ACTUAL, data_only=True, read_only=True)
    ws = wb["BASE_KSB1"]
    agregados = defaultdict(lambda: [0.0] * N_MESES_FECHADOS)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cc, classe, valor, mes, gestorial = row[2], row[3], row[16], row[18], row[19]
        if not isinstance(valor, (int, float)) or valor == 0:
            continue
        if not isinstance(mes, (int, float)) or not (1 <= mes <= N_MESES_FECHADOS):
            continue
        if not cc or gestorial is None:
            continue
        chave = (str(cc).strip(), str(gestorial).strip())
        agregados[chave][int(mes) - 1] += valor
    return agregados


def montar_proposta():
    efetivo_por_chave = carregar_efetivo_por_cc_gestorial()

    wb_origem = load_workbook(ORIGEM_FORECAST, data_only=True, read_only=True)
    ws_origem = wb_origem["DataBase_Detail"]

    header_classificacao = None
    linhas_saida = []
    n_com_efetivo = 0

    for i, row in enumerate(ws_origem.iter_rows(values_only=True)):
        if i == 1:
            header_classificacao = row[:N_COLS_CLASSIFICACAO]
            continue
        if i < 2:
            continue

        classificacao = list(row[:N_COLS_CLASSIFICACAO])
        metadados = [row[c] for c in COLS_METADADOS.values()]
        valor_mensal = [row[COL_VALOR_MENSAL_INICIO + m] for m in range(12)]
        reajuste = [row[COL_REAJUSTE_INICIO + m] for m in range(12)]
        valor_final = [row[COL_VALOR_FINAL_INICIO + m] for m in range(12)]

        if all(v is None for v in valor_final):
            continue

        cc_linha = str(row[1]).strip() if row[1] else None
        gestorial_linha = str(row[2]).strip() if row[2] is not None else None
        chave = (cc_linha, gestorial_linha)
        efetivo = efetivo_por_chave.get(chave, [0.0] * N_MESES_FECHADOS) + [None] * (12 - N_MESES_FECHADOS)
        if chave in efetivo_por_chave:
            n_com_efetivo += 1

        linhas_saida.append((classificacao, metadados, valor_mensal, reajuste, valor_final, efetivo))

    wb = Workbook()
    ws = wb.active
    ws.title = "Proposta Layout"

    # ---- linha 1: titulos de secao (mesclados) ----
    col = 1
    secoes = [
        ("CLASSIFICAÇÃO", N_COLS_CLASSIFICACAO + len(COLS_METADADOS), COR_HEADER_CLASSIFICACAO),
        ("FORECAST R7 2026", 36, COR_HEADER_FORECAST),
        ("EFETIVO (ACTUAL)", 13, COR_HEADER_EFETIVO),
        ("DIFERENÇA", 1, COR_HEADER_DIFERENCA),
    ]
    for titulo, largura, cor in secoes:
        c0 = col
        c1 = col + largura - 1
        ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
        cel = ws.cell(row=1, column=c0, value=titulo)
        cel.font = FONTE_SECAO
        cel.fill = PatternFill("solid", fgColor=cor)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(c0, c1 + 1):
            ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=cor)
        col += largura

    # ---- linha 2: cabecalho de coluna ----
    header = list(header_classificacao) + list(COLS_METADADOS.keys())
    for metrica in ["Valor Mensal", "% Reajuste", "Valor Final Previsão"]:
        for mes in MESES:
            header.append(f"{metrica} {mes}")
    for mes in MESES:
        header.append(f"Efetivo {mes}")
    header.append("TOTAL Efetivo (meses fechados)")
    header.append("Diferença (Forecast Total - Efetivo Total)")
    for j, val in enumerate(header, start=1):
        ws.cell(row=2, column=j, value=val)

    col_classif = 1
    col_metadados = col_classif + N_COLS_CLASSIFICACAO
    col_forecast = col_metadados + len(COLS_METADADOS)
    col_efetivo = col_forecast + 36
    col_total_efetivo = col_efetivo + 12
    col_diferenca = col_total_efetivo + 1

    for j in range(1, col_diferenca + 1):
        cel = ws.cell(row=2, column=j)
        cel.font = FONTE_HEADER
        if j < col_forecast:
            cel.fill = PatternFill("solid", fgColor=COR_HEADER_CLASSIFICACAO)
        elif j < col_efetivo:
            cel.fill = PatternFill("solid", fgColor=COR_HEADER_FORECAST)
        elif j < col_diferenca:
            cel.fill = PatternFill("solid", fgColor=COR_HEADER_EFETIVO)
        else:
            cel.fill = PatternFill("solid", fgColor=COR_HEADER_DIFERENCA)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ---- linhas de dado (indice explicito - linha 1 = titulos, linha 2 = header) ----
    for offset, (classificacao, metadados, valor_mensal, reajuste, valor_final, efetivo) in enumerate(linhas_saida):
        row_idx = offset + 3
        total_forecast = sum(v for v in valor_final if isinstance(v, (int, float)))
        total_efetivo = sum(v for v in efetivo if isinstance(v, (int, float)))
        diferenca = total_forecast - total_efetivo
        linha = (
            list(classificacao) + list(metadados)
            + list(valor_mensal) + list(reajuste) + list(valor_final)
            + list(efetivo) + [total_efetivo, diferenca]
        )
        for j, val in enumerate(linha, start=1):
            ws.cell(row=row_idx, column=j, value=val)

    n_linhas = len(linhas_saida) + 2  # +2 = linha 1 (titulos) + linha 2 (header)

    for row_idx in range(3, n_linhas + 1):
        for j in range(1, col_forecast):
            ws.cell(row=row_idx, column=j).fill = PatternFill("solid", fgColor=COR_CEL_CLASSIFICACAO)
        for j in range(col_forecast, col_efetivo):
            ws.cell(row=row_idx, column=j).fill = PatternFill("solid", fgColor=COR_CEL_FORECAST)
        for j in range(col_efetivo, col_diferenca):
            ws.cell(row=row_idx, column=j).fill = PatternFill("solid", fgColor=COR_CEL_EFETIVO)
        ws.cell(row=row_idx, column=col_diferenca).fill = PatternFill("solid", fgColor=COR_CEL_DIFERENCA)
        ws.cell(row=row_idx, column=col_diferenca).font = Font(bold=True)

        for j in range(col_forecast, col_forecast + 12):
            ws.cell(row=row_idx, column=j).number_format = "#,##0"
        for j in range(col_forecast + 12, col_forecast + 24):
            ws.cell(row=row_idx, column=j).number_format = "0.00%"
        for j in range(col_forecast + 24, col_efetivo):
            ws.cell(row=row_idx, column=j).number_format = "#,##0"
        for j in range(col_efetivo, col_diferenca + 1):
            ws.cell(row=row_idx, column=j).number_format = "#,##0"

    ws.freeze_panes = f"{get_column_letter(col_forecast)}3"

    for c in range(1, col_diferenca + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.column_dimensions[get_column_letter(8)].width = 40   # DETALHE SERVIÇO/PRODUTO
    ws.column_dimensions[get_column_letter(10)].width = 30  # Nome Fornecedor
    ws.row_dimensions[2].height = 32

    tabela = Table(displayName="TabelaDespesas", ref=f"A2:{get_column_letter(col_diferenca)}{n_linhas}")
    tabela.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
    ws.add_table(tabela)

    # ---- aba de notas ----
    ws_notas = wb.create_sheet("Notas do Layout")
    notas = [
        ("O que mudou", "Layout novo: Classificação (A-X, intacta) + Forecast (Valor Mensal/% Reajuste/Valor Final) + Efetivo (Actual) + Diferença. Antes eram 6 blocos repetidos sem nome claro."),
        ("Fonte do Forecast", "Detalhe_Despesas_Fitted Units_Forecast July.xlsx, aba DataBase_Detail (R7)."),
        ("Fonte do Efetivo", "KSB1 July Actual 2026.xlsx, aba BASE_KSB1 (oficial, cumulativa Jan-Jul). Pra atualizar mês a mês, troca pro arquivo Actual do mês mais recente fechado - a aba já vem cumulativa, não precisa somar vários arquivos."),
        ("Chave de casamento Forecast x Efetivo", "(Centro de Custo, Gestorial) - a melhor testada, mas só bate 69% das combinações (114 de 166). Ainda precisa ser refinada - combinado com a usuária que isso é uma v1, pra melhorar juntos depois."),
        ("Limitação importante", "O Efetivo é agregado por (CC, Gestorial), não por linha/fornecedor - se várias linhas do Forecast compartilham a mesma combinação, o mesmo valor de Efetivo aparece repetido em todas. NÃO somar a coluna Efetivo direto (duplica) - usar linha a linha ou agrupar por (CC, Gestorial) antes de somar."),
        (f"Cobertura", f"{n_com_efetivo} de {len(linhas_saida)} linhas do Forecast encontraram Efetivo correspondente."),
    ]
    ws_notas.append(["Tópico", "Detalhe"])
    for cel in ws_notas[1]:
        cel.font = Font(bold=True)
    for topico, detalhe in notas:
        ws_notas.append([topico, detalhe])
    ws_notas.column_dimensions["A"].width = 30
    ws_notas.column_dimensions["B"].width = 100
    for row in ws_notas.iter_rows(min_row=2):
        for cel in row:
            cel.alignment = Alignment(wrap_text=True, vertical="top")

    SAIDA.mkdir(parents=True, exist_ok=True)
    nome_arquivo = nome_com_versao(SAIDA, "MP2027_PROPOSTA_Layout_Detalhe_Despesas.xlsx")
    caminho_final = SAIDA / nome_arquivo
    wb.save(caminho_final)
    return caminho_final, len(linhas_saida), n_com_efetivo


if __name__ == "__main__":
    caminho, n, n_com_efetivo = montar_proposta()
    print(f"Proposta gerada: {caminho}")
    print(f"{n} linhas | {n_com_efetivo} com Efetivo encontrado ({n_com_efetivo/n:.0%})")
