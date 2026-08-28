"""MP2027 - despesas EFETIVAS Jan-Jul/2026, por unidade (Centro de Montagem)
x Fornecedor x Item de Compra, a partir do arquivo oficial e fechado
"KSB1 July Actual 2026.xlsx" (aba BASE_KSB1, cumulativa Jan-Jul).

Pedido da usuária (2026-08-28, revisão do primeiro arquivo por Gestorial -
"não gostei do que você fez"): CM | Fornecedor | Item de Compra | valor por
mês, uma aba por unidade, só DESPESA (exclui Mão de Obra - coluna DG/MO da
base, confirmado com a usuária que "MO" = rateio de folha/HR, sem
fornecedor real, deve ficar de fora). "Item de compra" = coluna "Texto do
pedido" (confirmado com a usuária - é o campo com a descrição real do que
foi comprado, ex: "Cremalheira Barra Industrial PPA"; "Denominação" fica
vazia nessas linhas).
"""
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "_shared"))
from ksb1_core import nome_com_versao  # noqa: E402

ORIGEM = Path(r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Custos Fitted Units\Resultados Fitted\2026\07 - Jul\07_Jul_Actual\KSB1 July Actual 2026.xlsx")
SAIDA = Path(r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Management Plan\MP2027")

MESES_LABEL = ["JAN", "FEV", "MAR", "ABR", "MAY", "JUN", "JUL"]
COL_FORNECEDOR_COD = 5
COL_FORNECEDOR_NOME = 6
COL_TEXTO_PEDIDO = 7
COL_DENOMINACAO = 8
COL_VALOR = 16
COL_MES = 18
COL_DESC_GESTORIAL = 20
COL_CM = 21
COL_VARIABILIDADE = 22
COL_DG_MO = 28

UNIDADES_ATIVAS = ["SJP", "IBI", "GOI", "RES", "GER"]


def processar():
    agregados = defaultdict(lambda: [0.0] * 7)  # (cm, fornecedor, item) -> meses
    fora_do_escopo = defaultdict(float)  # CM fora das 5 ativas (encerradas)
    sem_fornecedor = defaultdict(float)  # (cm) -> total, sem fornecedor identificado (nome E código vazios)
    total_mao_de_obra = 0.0

    wb = load_workbook(ORIGEM, data_only=True, read_only=True)
    ws = wb["BASE_KSB1"]

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        mes = row[COL_MES]
        cm = row[COL_CM]
        valor = row[COL_VALOR]
        dg_mo = row[COL_DG_MO]

        if not isinstance(valor, (int, float)) or valor == 0:
            continue
        if not isinstance(mes, (int, float)) or not (1 <= mes <= 7):
            continue
        if not cm:
            continue

        if dg_mo == "MO":
            total_mao_de_obra += valor
            continue

        if cm not in UNIDADES_ATIVAS:
            fora_do_escopo[cm] += valor
            continue

        nome_fornecedor = row[COL_FORNECEDOR_NOME]
        cod_fornecedor = row[COL_FORNECEDOR_COD]
        if nome_fornecedor and str(nome_fornecedor).strip():
            fornecedor = str(nome_fornecedor).strip()
        elif cod_fornecedor:
            fornecedor = f"Cód. {cod_fornecedor}"
        else:
            # nem nome nem codigo - nao e' fornecedor de verdade (lancamento
            # automatico de rateio, ex: PIS/COFINS, usuario sistemico "PROD")
            # - fica de fora das abas de fornecedor/item, achado 2026-08-28
            # ao investigar o volume de linhas do IBI (94% caia aqui).
            sem_fornecedor[cm] += valor
            continue

        item = row[COL_TEXTO_PEDIDO] or row[COL_DENOMINACAO]
        item = str(item).strip() if item else "(sem descrição)"
        desc_gestorial = row[COL_DESC_GESTORIAL] or ""
        variabilidade = row[COL_VARIABILIDADE] or ""

        chave = (cm, desc_gestorial, variabilidade, fornecedor, item)
        agregados[chave][int(mes) - 1] += valor

    return agregados, fora_do_escopo, sem_fornecedor, total_mao_de_obra


def gerar_excel(agregados, fora_do_escopo, sem_fornecedor, total_mao_de_obra) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    por_unidade = defaultdict(list)
    for (cm, desc_gestorial, variabilidade, fornecedor, item), meses in agregados.items():
        por_unidade[cm].append((desc_gestorial, variabilidade, fornecedor, item, meses))

    for cm in UNIDADES_ATIVAS:
        ws = wb.create_sheet(cm)
        ws.append(["CM", "Descrição Gestorial", "Variabilidade", "Fornecedor", "Item de Compra", *MESES_LABEL, "TOTAL"])
        for cel in ws[1]:
            cel.font = Font(bold=True)
        ws.freeze_panes = "A2"

        linhas = sorted(por_unidade.get(cm, []), key=lambda x: sum(x[4]), reverse=True)
        for desc_gestorial, variabilidade, fornecedor, item, meses in linhas:
            total = sum(meses)
            ws.append([cm, desc_gestorial, variabilidade, fornecedor, item, *[round(v, 2) for v in meses], round(total, 2)])

        for col_idx in range(6, 14):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"
        larguras = {1: 6, 2: 32, 3: 14, 4: 42, 5: 50, 13: 14}
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(col_idx, 11)

    ws_diag = wb.create_sheet("Fora do escopo")
    ws_diag.append(["O que", "Total Jan-Jul (R$)", "Observação"])
    for cel in ws_diag[1]:
        cel.font = Font(bold=True)
    ws_diag.append(["Mão de Obra (DG/MO = MO)", round(total_mao_de_obra, 2), "Rateio de folha/HR, sem fornecedor - excluído a pedido da usuária, só despesa entra"])
    for cm, total in sorted(sem_fornecedor.items(), key=lambda x: -abs(x[1])):
        ws_diag.append([f"Sem fornecedor identificado - unidade '{cm}'", round(total, 2), "Lançamento automático de rateio (ex: PIS/COFINS), sem nome nem código de fornecedor - não é compra de verdade"])
    for cm, total in sorted(fora_do_escopo.items(), key=lambda x: -abs(x[1])):
        ws_diag.append([f"Unidade '{cm}'", round(total, 2), "Encerrada ou fora das 5 unidades ativas - não entra no MP2027"])
    ws_diag.column_dimensions["A"].width = 30
    ws_diag.column_dimensions["C"].width = 70

    SAIDA.mkdir(parents=True, exist_ok=True)
    nome_arquivo = nome_com_versao(SAIDA, "MP2027_Despesas_por_Fornecedor_Item_Jan_Jul26.xlsx")
    caminho_final = SAIDA / nome_arquivo
    wb.save(caminho_final)
    return caminho_final


if __name__ == "__main__":
    agregados, fora_do_escopo, sem_fornecedor, total_mao_de_obra = processar()
    caminho = gerar_excel(agregados, fora_do_escopo, sem_fornecedor, total_mao_de_obra)

    totais_cm = defaultdict(float)
    n_linhas_cm = defaultdict(int)
    for (cm, *_), meses in agregados.items():
        totais_cm[cm] += sum(meses)
        n_linhas_cm[cm] += 1

    print(f"Arquivo gerado: {caminho}")
    print("\nTotal de DESPESA c/ fornecedor identificado, por unidade (Jan-Jul, R$) e nº de linhas (fornecedor x item):")
    for cm in UNIDADES_ATIVAS:
        print(f"  {cm}: R$ {totais_cm.get(cm, 0):,.0f}  ({n_linhas_cm.get(cm, 0)} linhas)")
    print(f"\nMão de Obra excluída (Jan-Jul, todas unidades): R$ {total_mao_de_obra:,.0f}")
    print("Sem fornecedor identificado (lançamento automático, ex: PIS/COFINS):")
    for cm, total in sorted(sem_fornecedor.items(), key=lambda x: -abs(x[1])):
        print(f"  {cm}: R$ {total:,.0f}")
    print("Fora do escopo (unidades encerradas):")
    for cm, total in sorted(fora_do_escopo.items(), key=lambda x: -abs(x[1])):
        print(f"  {cm}: R$ {total:,.0f}")
