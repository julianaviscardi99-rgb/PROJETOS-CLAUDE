"""MP2027 - Passo 0 (scoping): lista de principais fornecedores da Fitted Units,
com valor mensal, a partir do "Detalhe_Despesas_Fitted Units_Forecast" (aba
DataBase_Detail). Ponto de partida pra discutir o budget de plano - o
arquivo de origem tem "Nome Fornecedor" com variacoes de escrita pro mesmo
fornecedor, entao o agrupamento usa "Codigo Fornecedor" (estavel, do SAP) e
no output mostra todos os nomes distintos encontrados pra cada codigo, pra
dar visibilidade da bagunca em vez de esconder ela.

Valor mensal usado = colunas 94-106 ("R07 JAN".."R07 DEC" + "R07 TOTAL ANO"),
o 6o (e ultimo) bloco de colunas repetidas na aba - confirmado por
comparacao contra o total oficial da aba "Resumo Custos" (mesma forma da
curva mes a mes; diferenca residual de ~2-3%, provavelmente alguma regra de
exclusao/rateio que nao foi investigada nesta primeira passada).
"""
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "_shared"))
from ksb1_core import nome_com_versao  # noqa: E402

MESES = ["JAN", "FEV", "MAR", "ABR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
COL_CODIGO = 8
COL_NOME = 9
COL_VALOR_INICIO = 94  # bloco 6 (0-indexed): R07 JAN .. R07 DEC + R07 TOTAL ANO


def ler_fornecedores(caminho_origem: Path) -> tuple[list[dict], list[dict]]:
    """Devolve (fornecedores_identificados, linhas_sem_fornecedor).
    Linhas sem 'Código Fornecedor' preenchido na planilha de origem (gap de
    preenchimento, não custo interno de verdade - conferido manualmente:
    maioria é 'Outras Despesas'/'Aluguéis'/'Prestados - Serviços' sem nome
    de fornecedor também) vão separadas, agrupadas por categoria de custo
    (Descrição Gestorial), pra não distorcer o ranking de fornecedores."""
    wb = load_workbook(caminho_origem, data_only=True, read_only=True)
    ws = wb["DataBase_Detail"]

    valores = defaultdict(lambda: [0.0] * 12)
    nomes_vistos = defaultdict(Counter)
    sem_fornecedor = defaultdict(lambda: [0.0] * 12)

    # rótulos que aparecem no campo "Nome Fornecedor" mas não são fornecedor
    # de verdade (lançamento contábil/operação interna) - achado ao inspecionar
    # o output ("Baixa de Materiais" = baixa de estoque, não uma empresa).
    NOMES_INVALIDOS = {"baixa de materiais"}

    def vazio(v) -> bool:
        if v is None:
            return True
        texto = str(v).strip()
        return texto in ("", "-") or texto.lower() in NOMES_INVALIDOS

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        codigo = row[COL_CODIGO]
        nome = row[COL_NOME]
        meses_linha = [row[COL_VALOR_INICIO + m] for m in range(12)]

        if vazio(codigo):
            if not vazio(nome):
                # sem código, mas com nome real (ex: "Unimed", "Flash") -
                # agrupa pelo nome mesmo, pra não perder o fornecedor
                chave = f"NOME::{str(nome).strip()}"
                nomes_vistos[chave][str(nome).strip()] += 1
            else:
                categoria = row[3] or "(sem categoria)"
                for m, v in enumerate(meses_linha):
                    if isinstance(v, (int, float)):
                        sem_fornecedor[categoria][m] += v
                continue
        else:
            chave = str(codigo).strip()
            if not vazio(nome):
                nomes_vistos[chave][str(nome).strip()] += 1

        for m, v in enumerate(meses_linha):
            if isinstance(v, (int, float)):
                valores[chave][m] += v

    linhas = []
    for chave, meses in valores.items():
        contagem_nomes = nomes_vistos[chave]
        nome_principal = contagem_nomes.most_common(1)[0][0] if contagem_nomes else "(sem nome)"
        outros_nomes = sorted(n for n in contagem_nomes if n != nome_principal)
        linhas.append({
            "codigo": "(sem código)" if chave.startswith("NOME::") else chave,
            "nome": nome_principal,
            "outros_nomes": "; ".join(outros_nomes),
            "meses": meses,
            "total": sum(meses),
        })
    linhas.sort(key=lambda x: x["total"], reverse=True)

    linhas_sem_fornecedor = [
        {"categoria": categoria, "meses": meses, "total": sum(meses)}
        for categoria, meses in sem_fornecedor.items()
    ]
    linhas_sem_fornecedor.sort(key=lambda x: x["total"], reverse=True)

    return linhas, linhas_sem_fornecedor


def _formatar_aba(ws, n_col_meses_inicio: int, n_col_meses_fim: int):
    for cel in ws[1]:
        cel.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col_idx in range(n_col_meses_inicio, n_col_meses_fim + 1):
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"


def gerar_excel(linhas: list[dict], linhas_sem_fornecedor: list[dict], pasta_saida: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Fornecedores"
    ws.append(["Código Fornecedor", "Nome Fornecedor", *MESES, "TOTAL ANO", "Outros nomes encontrados p/ este código"])
    for linha in linhas:
        ws.append([linha["codigo"], linha["nome"], *[round(v, 2) for v in linha["meses"]], round(linha["total"], 2), linha["outros_nomes"]])
    _formatar_aba(ws, 3, 15)
    larguras = {1: 16, 2: 42, 15: 14, 16: 50}
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(col_idx, 11)

    ws2 = wb.create_sheet("Sem Fornecedor Identificado")
    ws2.append(["Categoria de Custo (Descrição Gestorial)", *MESES, "TOTAL ANO"])
    for linha in linhas_sem_fornecedor:
        ws2.append([linha["categoria"], *[round(v, 2) for v in linha["meses"]], round(linha["total"], 2)])
    _formatar_aba(ws2, 2, 14)
    larguras2 = {1: 42, 14: 14}
    for col_idx in range(1, ws2.max_column + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = larguras2.get(col_idx, 11)

    pasta_saida.mkdir(parents=True, exist_ok=True)
    nome_arquivo = nome_com_versao(pasta_saida, "MP2027_Lista_Fornecedores_base_Forecast_Jul26.xlsx")
    caminho_final = pasta_saida / nome_arquivo
    wb.save(caminho_final)
    return caminho_final


if __name__ == "__main__":
    ORIGEM = Path(r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Custos Fitted Units\Resultados Fitted\2026\07 - Jul\07_Jul_Forecast\Detalhe_Despesas_Fitted Units_Forecast July.xlsx")
    SAIDA = Path(r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Management Plan\MP2027")

    linhas, linhas_sem_fornecedor = ler_fornecedores(ORIGEM)
    caminho = gerar_excel(linhas, linhas_sem_fornecedor, SAIDA)

    total_sem_fornecedor = sum(l["total"] for l in linhas_sem_fornecedor)
    print(f"{len(linhas)} fornecedores (por código) encontrados.")
    print(f"R$ {total_sem_fornecedor:,.0f} sem fornecedor identificado na planilha (aba separada, {len(linhas_sem_fornecedor)} categorias de custo).")
    print(f"Arquivo gerado: {caminho}\n")
    print(f"{'Fornecedor':45} {'Código':14} {'TOTAL ANO':>14}")
    for linha in linhas[:20]:
        print(f"{linha['nome'][:45]:45} {linha['codigo']:14} {linha['total']:>14,.0f}")
