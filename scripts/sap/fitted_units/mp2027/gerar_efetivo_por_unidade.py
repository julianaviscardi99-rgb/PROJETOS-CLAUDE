"""MP2027 - custo EFETIVO Jan-Jul/2026 por unidade, a partir do arquivo
oficial e fechado "KSB1 July Actual 2026.xlsx" (rede, pasta 07_Jul_Actual)
- aba BASE_KSB1, que é CUMULATIVA (acumula todas as linhas de Jan a Jul do
ano, não só julho) e já vem com Gestorial/Centro de Montagem RESOLVIDOS por
fórmula (colunas 19/21), a mesma classificação usada no fechamento oficial
(Base Intermediária/EBIT) - não precisa de mapeamento próprio.

Pedido da usuária (2026-08-28): CM | Gestorial | Fornecedor | valor por mês
(Jan-Jul), uma aba por unidade, pra debater budget de plano MP2027 com os
gerentes. Substitui uma primeira tentativa que usava um extrato bruto
"Sem Agrupamento" à parte (data/processed/energia_eletrica_fitted/...) -
a usuária pediu pra usar este arquivo oficial em vez disso.
"""
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "_shared"))
from ksb1_core import nome_com_versao  # noqa: E402

ORIGEM = Path(r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Custos Fitted Units\Resultados Fitted\2026\07 - Jul\07_Jul_Actual\KSB1 July Actual 2026.xlsx")
SAIDA = Path(r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Management Plan\MP2027")

MESES_LABEL = ["JAN", "FEV", "MAR", "ABR", "MAY", "JUN", "JUL"]
COL_FORNECEDOR_COD = 5
COL_FORNECEDOR_NOME = 6
COL_VALOR = 16
COL_MES = 18
COL_GESTORIAL = 19
COL_DESC_GESTORIAL = 20
COL_CM = 21

UNIDADES_ATIVAS = ["SJP", "IBI", "GOI", "RES", "GER"]


def processar():
    agregados = defaultdict(lambda: [0.0] * 7)  # (cm, gestorial, desc, fornecedor) -> meses
    fora_do_escopo = defaultdict(float)  # cm que não está em UNIDADES_ATIVAS (encerradas/residual)

    wb = load_workbook(ORIGEM, data_only=True, read_only=True)
    ws = wb["BASE_KSB1"]

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        mes = row[COL_MES]
        cm = row[COL_CM]
        valor = row[COL_VALOR]
        if not isinstance(valor, (int, float)) or valor == 0:
            continue
        if not isinstance(mes, (int, float)) or not (1 <= mes <= 7):
            continue
        if not cm:
            continue

        gestorial = row[COL_GESTORIAL]
        desc_gestorial = row[COL_DESC_GESTORIAL] or ""
        nome_fornecedor = row[COL_FORNECEDOR_NOME]
        cod_fornecedor = row[COL_FORNECEDOR_COD]

        if nome_fornecedor and str(nome_fornecedor).strip():
            fornecedor = str(nome_fornecedor).strip()
        elif cod_fornecedor:
            fornecedor = f"Cód. {cod_fornecedor}"
        else:
            fornecedor = "(sem fornecedor)"

        if cm not in UNIDADES_ATIVAS:
            fora_do_escopo[cm] += valor
            continue

        chave = (cm, str(gestorial) if gestorial is not None else "", desc_gestorial, fornecedor)
        agregados[chave][int(mes) - 1] += valor

    return agregados, fora_do_escopo


def gerar_excel(agregados, fora_do_escopo) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    por_unidade = defaultdict(list)
    for (cm, gest_cod, gest_desc, fornecedor), meses in agregados.items():
        por_unidade[cm].append((gest_cod, gest_desc, fornecedor, meses))

    for cm in UNIDADES_ATIVAS:
        ws = wb.create_sheet(cm)
        ws.append(["CM", "Gestorial", "Descrição Gestorial", "Fornecedor", *MESES_LABEL, "TOTAL"])
        for cel in ws[1]:
            cel.font = Font(bold=True)
        ws.freeze_panes = "A2"

        linhas = sorted(por_unidade.get(cm, []), key=lambda x: sum(x[3]), reverse=True)
        for gest_cod, gest_desc, fornecedor, meses in linhas:
            total = sum(meses)
            ws.append([cm, gest_cod, gest_desc, fornecedor, *[round(v, 2) for v in meses], round(total, 2)])

        for col_idx in range(5, 13):
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0"
        larguras = {1: 6, 2: 14, 3: 40, 4: 42, 12: 14}
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = larguras.get(col_idx, 11)

    ws_diag = wb.create_sheet("Fora do escopo (residual)")
    ws_diag.append(["Centro de Montagem", "Total Jan-Jul (R$)", "Observação"])
    for cel in ws_diag[1]:
        cel.font = Font(bold=True)
    for cm, total in sorted(fora_do_escopo.items(), key=lambda x: -abs(x[1])):
        ws_diag.append([cm, round(total, 2), "Unidade encerrada ou fora das 5 unidades ativas - não entra no MP2027"])
    ws_diag.column_dimensions["A"].width = 20
    ws_diag.column_dimensions["C"].width = 55

    SAIDA.mkdir(parents=True, exist_ok=True)
    nome_arquivo = nome_com_versao(SAIDA, "MP2027_Efetivo_por_Unidade_Jan_Jul26.xlsx")
    caminho_final = SAIDA / nome_arquivo
    wb.save(caminho_final)
    return caminho_final


if __name__ == "__main__":
    agregados, fora_do_escopo = processar()
    caminho = gerar_excel(agregados, fora_do_escopo)

    totais_cm = defaultdict(float)
    for (cm, *_), meses in agregados.items():
        totais_cm[cm] += sum(meses)

    print(f"Arquivo gerado: {caminho}")
    print("\nTotal por unidade (Jan-Jul, R$):")
    for cm in UNIDADES_ATIVAS:
        print(f"  {cm}: {totais_cm.get(cm, 0):,.0f}")
    print("\nFora do escopo (residual/encerradas):")
    for cm, total in sorted(fora_do_escopo.items(), key=lambda x: -abs(x[1])):
        print(f"  {cm}: {total:,.0f}")
