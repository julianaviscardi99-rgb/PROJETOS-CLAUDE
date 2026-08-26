#!/usr/bin/env python3
"""
Passo 5 (Rateio de Custos) do processo recorrente (Fitted Units Despesas).

Substitui o arquivo manual "_Abertura custos Fitted Units <Mes> <Ciclo>
<Ano>.xlsx" (link externo trocado a mao todo mes + celula de mes/Ciclo
trocada a mao) por um arquivo novo, gerado direto a partir da Base
Intermediaria do mes/Ciclo (ja gerada pelo Passo 4) - sem link nenhum.

Desenhado com a Juliana em 2026-08-25 (ver memory/BRIEFING.md e
memory/DECISOES.md dessa data pro historico completo da conversa):

1. Le a aba "Intermediaria" da Base Intermediaria do mes/Ciclo (mesma pra
   Actual e Flash, sem excecao) e classifica cada linha por Mini-Fabrica
   (unidade), usando EXCLUSIVAMENTE as colunas AA ("Var.") e AJ ("Conta
   Geral") - ja resolvidas pela propria Base Intermediaria. A coluna H
   ("Tp.Custo") NUNCA e' lida nem considerada pra classificacao (vem em
   branco pra algumas unidades, e mesmo onde vem preenchida nao e' a fonte
   de verdade - correcao explicita da usuaria, 2026-08-26). Mapeamento
   proprio por conta tambem foi abandonado (pedido explicito da usuaria,
   2026-08-25: "usa a base intermediaria... os valores tem que voltar
   exatamente" - ver _resolver_subcategoria).
2. Unidades ativas: SJP (0490), IBI (0491), GOI (0481), RES (0483).
   Gerencia (0499) e' tratada como uma "unidade" a parte (coluna GER no
   quadro sem rateio), nao recebe nem manda rateio pra si mesma. A
   Gerencia e' sempre 100% Fixa POR DEFINICAO DE NEGOCIO, nao por
   coincidencia: custo Variavel e' custo ligado a producao, e a Gerencia
   nao produz - logo nao deveria existir nenhum custo Variavel nela
   (correcao explicita da usuaria, 2026-08-26). Se a Base Intermediaria
   trouxer algo tipo='V' pra Gerencia mesmo assim (deveria ser impossivel),
   e' tratado como anomalia de lancamento e ignorado no rateio, nunca
   silenciosamente - ver _apenas_fixo.
3. Unidades ENCERRADAS (ontology/fitted_units.json ->
   centros_de_custo_por_unidade, status "encerrada", chave por Centro de
   Custo - mesma fonte que gerar_base_intermediaria.py ja usa) nunca entram
   no quadro nem recebem rateio (sem faturamento). Se aparecer residuo de
   custo nelas:
   - Sorocaba: fica de fora do rateio (esta em reclassificacao pra custo
     nao-recorrente, tratada em outro lugar) - so' aparece no aviso.
   - Qualquer outra unidade encerrada: o residuo soma no total da Gerencia
     antes de aplicar o rateio.
   Em ambos os casos, um aviso com as linhas (conta/descricao/unidade/valor)
   aparece no arquivo gerado - nunca fica escondido so' no log.
4. O rateio (% por unidade, muda geralmente em Janeiro mas pode mudar fora
   de epoca) fica num arquivo de configuracao separado
   (ontology/rateio_gerencia.json), NUNCA hardcoded aqui - ver
   carregar_rateio_vigente. O rateio e' espalhado CATEGORIA A CATEGORIA
   (mesma logica do arquivo real de Forecast, "Detalhe_Despesas_Fitted
   Units", aba "Resumo Custos") - cada linha (Labour, Depreciation etc.)
   recebe o pedaco dela da Gerencia, nao e' mais uma linha unica.
5. Gera 3 blocos no arquivo de saida: tabela do rateio vigente, quadro "sem
   rateio" (unidades ativas + GER, cada uma com seu proprio custo) e quadro
   "com rateio" (so' unidades ativas, ja' com o rateio espalhado por
   categoria, mais uma linha informativa "Rateio Gerencia" fora do Total
   Costs e um Check provando que o TOTAL dos dois quadros bate).
6. Salva no mesmo racional de sempre (resolver_pasta_ciclo), nunca
   sobrescreve (nome_com_versao).

NAO mexe em nada da Base Intermediaria - so' le (openpyxl, ReadOnly,
data_only=True - nao precisa de Excel/COM, os valores ja estao calculados
no arquivo).
"""
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "_shared"))
from ksb1_core import (  # noqa: E402
    MESES_PASTA,
    REDE_BASE,
    encontrar_arquivo_mais_recente,
    nome_com_versao,
    resolver_pasta_ciclo,
)

MESES_INGLES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}
MESES_ABREV = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

ONTOLOGY_DIR = Path(__file__).resolve().parents[4] / "ontology"
ONTOLOGY_UNIDADES_PATH = ONTOLOGY_DIR / "fitted_units.json"
RATEIO_CONFIG_PATH = ONTOLOGY_DIR / "rateio_gerencia.json"

# Mini-Fabrica -> sigla, so' as unidades ATIVAS que entram no quadro/rateio.
# Confirmado com a usuaria em 2026-08-25 (RES ainda sem custo real ate essa
# data, codigo confirmado por ela: "0483 e' de Resende").
UNIDADES_ATIVAS = {
    "0490": "SJP",
    "0491": "IBI",
    "0481": "GOI",
    "0483": "RES",
}
GERENCIA_MINIFABRICA = "0499"
SIGLA_GERENCIA = "GER"

# Ordem de exibicao das unidades no quadro (GER so' aparece no quadro "sem
# rateio", nunca no "com rateio").
ORDEM_UNIDADES_ATIVAS = ["SJP", "IBI", "GOI", "RES"]

ORDEM_VARIAVEL = ["Labour", "Handling", "Direct Materials", "Transportation", "Other Variable"]
ORDEM_FIXO = ["Labour", "Depreciation", "IFRS16", "Rents", "Condominio", "Other Fixed"]

# A Base Intermediária (aba Intermediária) já traz, em colunas próprias,
# a classificação RESOLVIDA de cada linha. Usamos EXCLUSIVAMENTE AA ("Var.")
# e AJ ("Conta Geral") - a coluna H ("Tp.Custo") nunca é lida (correção
# explícita da usuária, 2026-08-26: não é só que H "vem em branco às vezes",
# é que H nunca deve ser considerado pra variabilidade, ponto final). Usar
# essas colunas prontas (em vez de reclassificar por conta na mão, que já
# deu resultado errado pra Handling/Rents num teste anterior) é o que faz
# os valores baterem de verdade.
COL_VAR = 27          # AA - 'F' ou 'V', já resolvido linha a linha
COL_CONTA_GERAL = 36  # AJ - subcategoria (Labour, Handling, Depreciations,
                      # IFRS16 (Amortization), Rents, Transport, Others,
                      # Prod.Consumables)

# Conta Geral (coluna AJ) -> nosso rótulo de exibição (o mockup da usuária
# usa nomes um pouco diferentes do arquivo). "Others" se resolve por tipo
# (Other Variable/Other Fixed) dentro de _resolver_subcategoria.
_CONTA_GERAL_PARA_SUBCATEGORIA = {
    "Labour": "Labour",
    "Handling": "Handling",
    "Prod.Consumables": "Direct Materials",
    "Transport": "Transportation",
    "Depreciations": "Depreciation",
    "IFRS16 (Amortization)": "IFRS16",
    "Rents": "Rents",
    "Condominio": "Condominio",
    "Condomínio": "Condominio",
}


def _resolver_subcategoria(tipo: str, conta_geral):
    """Traduz o valor da coluna AJ (Conta Geral) da Base Intermediária pro
    nosso rótulo de exibição, considerando se ele faz sentido dentro do
    macro certo (Variable Cost só tem Labour/Handling/Direct Materials/
    Transportation/Other Variable; Fixed Cost só tem Labour/Depreciation/
    IFRS16/Rents/Condominio/Other Fixed - ver ORDEM_VARIAVEL/ORDEM_FIXO).
    Devolve None se não reconhecer o valor ou ele não couber no macro desse
    tipo (ex: Base Intermediária às vezes marca 'Rents' como Variável, mas
    não existe linha "Rents" no Variable Cost do quadro) - quem chama cai
    no fallback (Other Variable/Fixed) nesse caso."""
    if conta_geral == "Others":
        return "Other Variable" if tipo == "V" else "Other Fixed"
    nome = _CONTA_GERAL_PARA_SUBCATEGORIA.get(conta_geral)
    if nome is None:
        return None
    ordem = ORDEM_VARIAVEL if tipo == "V" else ORDEM_FIXO
    return nome if nome in ordem else None


def carregar_centros_encerrados() -> dict:
    """Retorna {centro_de_custo (str): nome_da_unidade} pra toda unidade com
    status 'encerrada' - mesma fonte/formato que gerar_base_intermediaria.py
    ja usa (ontology/fitted_units.json), pra nao duplicar a lista."""
    dados = json.loads(ONTOLOGY_UNIDADES_PATH.read_text(encoding="utf-8"))
    grupos = dados["centros_de_custo_por_unidade"]["grupos"]
    mapa = {}
    for nome_grupo, info in grupos.items():
        if info.get("status") == "encerrada":
            for c in info["centros"]:
                mapa[str(c)] = nome_grupo
    return mapa


def carregar_rateio_vigente(mes: int, ano: int) -> tuple[dict, str]:
    """Le ontology/rateio_gerencia.json e devolve (percentuais, vigente_desde)
    da entrada mais recente cujo 'vigente_desde' (formato 'AAAA-MM') seja
    <= o mes/ano pedido. NUNCA hardcoda a %, porque ela muda (geralmente em
    Janeiro, mas as vezes fora de epoca - ex: quando a Resende entrou)."""
    dados = json.loads(RATEIO_CONFIG_PATH.read_text(encoding="utf-8"))
    alvo = f"{ano:04d}-{mes:02d}"
    candidatas = [e for e in dados["entradas"] if e["vigente_desde"] <= alvo]
    if not candidatas:
        raise RuntimeError(
            f"Não encontrei nenhuma % de rateio vigente pra {alvo} em "
            f"{RATEIO_CONFIG_PATH} — confira o arquivo de configuração."
        )
    escolhida = max(candidatas, key=lambda e: e["vigente_desde"])
    return escolhida["percentuais"], escolhida["vigente_desde"]


def localizar_base_intermediaria(mes: int, ano: int, ciclo: str) -> Path:
    pasta = resolver_pasta_ciclo(REDE_BASE / str(ano) / MESES_PASTA[mes], mes, ciclo)
    nome = f"Base Intermediária Fitted {MESES_INGLES[mes]} {ciclo} {ano}.xlsx"
    caminho = encontrar_arquivo_mais_recente(pasta, nome)
    if caminho is None:
        raise FileNotFoundError(
            f"Não encontrei '{nome}' em {pasta} — rode o Passo 4 (Base "
            "Intermediária) pra esse mês/Ciclo antes de rodar o Rateio de Custos."
        )
    return caminho


def _fmt_moeda(v):
    if v is None:
        v = 0
    v = round(v, 1)
    if v < 0:
        return f"({abs(v):,.1f})".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{v:,.1f}".replace(",", "X").replace(".", ",").replace("X", ".")


def ler_e_classificar(caminho_base_intermediaria: Path, mes: int, log):
    """Le a aba Intermediaria e devolve:
    - totais_ativos: {sigla_unidade: {(tipo, subcategoria): valor}}
      (SJP/IBI/GOI/RES/GER; chave especial "Não Classificado" quando a
      Base Intermediária não trouxe Var./Conta Geral pra aquela linha)
    - residuos_encerradas: lista de dicts (unidade, conta, descricao, valor)
      pra toda unidade encerrada com valor != 0 no mes
    - contas_nao_mapeadas: set de (tipo, conta_geral, desc) cujo valor da
      coluna Conta Geral não bateu com nenhuma subcategoria esperada pro
      tipo (caiu no fallback Other Variable/Fixed)
    - raw_por_mini_fabrica: {mini_fabrica (str): soma bruta de TODAS as
      linhas com esse código, independente de classificação/escopo} - usado
      pelo "check" por unidade (calcular_check_por_unidade), pra provar que
      nada se perde silenciosamente na classificação (pedido explícito da
      usuária, 2026-08-26).
    - fora_de_escopo: lista de dicts (conta, descricao, mini_fabrica,
      centro_custo, valor) pra linhas cujo Mini-Fábrica/Centro de Custo não
      bate com nenhuma unidade ativa, Gerência ou unidade encerrada
      conhecida - antes eram descartadas em silêncio, agora ficam
      registradas pra aparecer na aba "Comentários" do arquivo gerado.
    Valores em '000 BRL, custo positivo (mesma convencao do arquivo antigo:
    o valor bruto do SAP vem negativo pra custo, aqui ja inverte o sinal).

    Classificação Variável/Fixo e subcategoria vêm DIRETO e EXCLUSIVAMENTE
    das colunas AA ("Var.") e AJ ("Conta Geral") da própria Base
    Intermediária - a coluna H ("Tp.Custo") nunca é lida nem considerada
    (correção explícita da usuária, 2026-08-26)."""
    wb = __import__("openpyxl").load_workbook(
        caminho_base_intermediaria, read_only=True, data_only=True, keep_links=False
    )
    ws = wb["Intermediária"]
    col_mes = 8 + mes  # A=1...H=8, I(Jan)=9 -> mes=1 vira coluna 9

    centros_encerrados = carregar_centros_encerrados()

    totais_ativos = {sigla: {} for sigla in list(UNIDADES_ATIVAS.values()) + [SIGLA_GERENCIA]}
    residuos_encerradas = []
    contas_nao_mapeadas = set()
    raw_por_mini_fabrica = {}
    fora_de_escopo = []

    max_col = max(col_mes, COL_CONTA_GERAL)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col):
        conta = row[0].value  # A - Conta Gestorial
        desc_conta = row[1].value  # B
        centro_custo = row[4].value  # E
        mini_fabrica = row[5].value  # F
        tipo = row[COL_VAR - 1].value  # AA - 'F' ou 'V', ja resolvido
        conta_geral = row[COL_CONTA_GERAL - 1].value  # AJ - subcategoria
        if conta is None:
            continue
        valor_bruto = row[col_mes - 1].value if col_mes - 1 < len(row) else None
        if not valor_bruto:
            continue
        valor = -valor_bruto / 1000  # mesma convencao do arquivo antigo

        try:
            conta_int = int(conta)
        except (TypeError, ValueError):
            conta_int = conta

        # Classifica a chave (tipo, subcategoria) ANTES de decidir a unidade -
        # precisa dela tanto pra unidade ativa/Gerencia quanto pro residuo de
        # unidade encerrada (que, quando nao e' Sorocaba, entra no balde da
        # Gerencia NESSA MESMA categoria).
        if tipo not in ("V", "F"):
            chave = "Não Classificado"
        else:
            subcat = _resolver_subcategoria(tipo, conta_geral)
            if subcat is None:
                contas_nao_mapeadas.add((tipo, conta_geral, desc_conta))
                subcat = "Other Variable" if tipo == "V" else "Other Fixed"
            # Chave = (tipo, subcategoria), NUNCA so' a subcategoria: "Labour"
            # existe tanto em Variable Cost quanto em Fixed Cost - se a chave
            # fosse so' o nome, os dois se misturariam num balde so' e o
            # Total Costs contaria o mesmo valor 2x (bug real encontrado e
            # corrigido em 2026-08-25, testando contra Julho/Actual real).
            chave = (tipo, subcat)

        mini_fabrica_str = str(mini_fabrica).strip() if mini_fabrica is not None else ""
        centro_custo_str = str(centro_custo).strip() if centro_custo is not None else ""

        # Soma bruta por Mini-Fábrica, independente de classificação/escopo -
        # base do "check" por unidade (prova que o que entra no quadro bate
        # com o total real da Base Intermediária pra aquele código).
        if mini_fabrica_str:
            raw_por_mini_fabrica[mini_fabrica_str] = raw_por_mini_fabrica.get(mini_fabrica_str, 0) + valor

        if mini_fabrica_str in UNIDADES_ATIVAS:
            sigla = UNIDADES_ATIVAS[mini_fabrica_str]
        elif mini_fabrica_str == GERENCIA_MINIFABRICA:
            sigla = SIGLA_GERENCIA
        elif centro_custo_str in centros_encerrados:
            unidade_encerrada = centros_encerrados[centro_custo_str]
            residuos_encerradas.append(
                {
                    "unidade": unidade_encerrada,
                    "conta": conta_int,
                    "descricao": desc_conta,
                    "valor": valor,
                }
            )
            if unidade_encerrada.upper() != "SOROCABA":
                # Soma no balde da Gerencia NESSA MESMA categoria (tipo,
                # subcategoria) - assim o residuo entra no rateio por
                # categoria exatamente como se fosse custo proprio da
                # Gerencia (pedido explicito da usuaria, 2026-08-25).
                totais_ativos[SIGLA_GERENCIA][chave] = (
                    totais_ativos[SIGLA_GERENCIA].get(chave, 0) + valor
                )
            continue
        else:
            # Nem unidade ativa, nem Gerencia, nem unidade encerrada
            # conhecida (ex: FATURAMENTO, que ja e' pra ignorar) - fora de
            # escopo, nao entra no quadro nem no aviso principal, mas fica
            # registrada (fora_de_escopo) pra aparecer na aba "Comentários" -
            # antes era descartada em silêncio total (correção 2026-08-26).
            fora_de_escopo.append(
                {
                    "conta": conta_int,
                    "descricao": desc_conta,
                    "mini_fabrica": mini_fabrica_str,
                    "centro_custo": centro_custo_str,
                    "valor": valor,
                }
            )
            continue

        totais_ativos[sigla][chave] = totais_ativos[sigla].get(chave, 0) + valor

    if contas_nao_mapeadas:
        for tipo, conta_geral, desc in sorted(contas_nao_mapeadas, key=lambda x: (x[0], str(x[1]))):
            log(
                f"AVISO: '{desc}' (Conta Geral='{conta_geral}', tipo {tipo}) não bateu com "
                "nenhuma subcategoria esperada — caiu no 'Other Variable/Fixed' por padrão."
            )

    return totais_ativos, residuos_encerradas, contas_nao_mapeadas, raw_por_mini_fabrica, fora_de_escopo


def _apenas_fixo(dados_unidade: dict, log=None, sigla="") -> dict:
    """O custo da Gerência é 100% Fixo por definição de negócio (correção
    explícita da usuária, 2026-08-26): custo Variável é custo ligado à
    produção, e a Gerência não produz - então não deveria existir NENHUM
    custo Variável nela, nunca. Filtra o dicionário {(tipo, subcategoria):
    valor} mantendo só as chaves tipo='F'. Se aparecer algo tipo='V' com
    valor mesmo assim (não deveria ser possível, mas a Base Intermediária é
    a fonte de verdade e pode trazer isso por engano de lançamento), é
    tratado como anomalia: avisa e ignora - nunca vira rateio."""
    resultado = {}
    for chave, valor in dados_unidade.items():
        if isinstance(chave, tuple) and chave[0] == "V":
            if valor and log:
                log(
                    f"AVISO: {sigla} teve custo classificado como Variável "
                    f"({chave[1]}: {_fmt_moeda(valor)}) — a Gerência é sempre "
                    "Fixa, esse valor foi ignorado no rateio."
                )
            continue
        resultado[chave] = valor
    return resultado


def calcular_rateio(totais_ativos: dict, residuos_encerradas: list, percentuais: dict, log):
    """Devolve (rateio_por_unidade, gerencia_total, residuo_somado_gerencia,
    residuo_sorocaba_fora). `totais_ativos[GER]` ja' inclui o residuo de
    unidades encerradas (exceto Sorocaba) - foi somado categoria a categoria
    em `ler_e_classificar`. Aqui so' calcula o total (pra linha informativa
    "Rateio Gerência") e loga os avisos de residuo. `gerencia_total` conta
    so' o custo Fixo da Gerência (ver _apenas_fixo)."""
    ger = _apenas_fixo(totais_ativos.get(SIGLA_GERENCIA, {}), log, SIGLA_GERENCIA)
    gerencia_total = sum(ger.values())

    residuo_somado_gerencia = 0.0
    residuo_sorocaba_fora = 0.0
    for r in residuos_encerradas:
        if r["unidade"].upper() == "SOROCABA":
            residuo_sorocaba_fora += r["valor"]
        else:
            residuo_somado_gerencia += r["valor"]
            log(
                f"AVISO: resíduo de {r['unidade']} (conta {r['conta']} - {r['descricao']}: "
                f"{_fmt_moeda(r['valor'])}) somado ao custo da Gerência (mesma categoria) antes do rateio."
            )
    if residuo_sorocaba_fora:
        log(
            f"AVISO: resíduo de Sorocaba no mês ({_fmt_moeda(residuo_sorocaba_fora)}) — "
            "NÃO entra no rateio (unidade em reclassificação para custo não recorrente)."
        )

    # Linha informativa "Rateio Gerência" (fora do Total Costs - o rateio de
    # verdade ja' foi espalhado categoria a categoria em
    # calcular_dados_com_rateio): quanto cada unidade recebeu no total.
    rateio_por_unidade = {}
    for sigla in ORDEM_UNIDADES_ATIVAS:
        pct = percentuais.get(sigla, 0)
        rateio_por_unidade[sigla] = gerencia_total * pct

    return rateio_por_unidade, gerencia_total, residuo_somado_gerencia, residuo_sorocaba_fora


def calcular_check_por_unidade(totais_ativos: dict, raw_por_mini_fabrica: dict, residuos_encerradas: list) -> dict:
    """Confere, unidade por unidade, se o total que entrou no quadro "sem
    rateio" bate com o total bruto da Base Intermediária pra aquele
    Mini-Fábrica - pedido explícito da usuária, 2026-08-26: um check logo
    abaixo do Total Costs de cada unidade, provando que nada se perdeu
    silenciosamente na classificação. Devolve {sigla: (diff, esperado,
    classificado)}.

    Pras unidades ativas, `esperado` é só a soma bruta do próprio código de
    Mini-Fábrica - deveria SEMPRE bater exato (diff=0), porque toda linha
    com esse código vai pra `totais_ativos[sigla]` sem exceção. Pra
    Gerência, `esperado` soma o próprio código (0499) MAIS o resíduo de
    unidades encerradas (exceto Sorocaba) - que por regra de negócio entra
    no balde da Gerência antes do rateio (ver ler_e_classificar) - também
    deveria bater exato."""
    codigo_por_sigla = {v: k for k, v in UNIDADES_ATIVAS.items()}
    residuo_gerencia = sum(
        r["valor"] for r in residuos_encerradas if r["unidade"].upper() != "SOROCABA"
    )
    checks = {}
    for sigla in ORDEM_UNIDADES_ATIVAS:
        codigo = codigo_por_sigla[sigla]
        esperado = raw_por_mini_fabrica.get(codigo, 0)
        classificado = sum(totais_ativos.get(sigla, {}).values())
        checks[sigla] = (classificado - esperado, esperado, classificado)
    esperado_ger = raw_por_mini_fabrica.get(GERENCIA_MINIFABRICA, 0) + residuo_gerencia
    classificado_ger = sum(totais_ativos.get(SIGLA_GERENCIA, {}).values())
    checks[SIGLA_GERENCIA] = (classificado_ger - esperado_ger, esperado_ger, classificado_ger)
    return checks


def calcular_dados_com_rateio(totais_ativos: dict, percentuais: dict, log=None) -> dict:
    """Devolve {unidade: {(tipo, subcategoria): valor}} pras unidades ATIVAS,
    com o custo da Gerência ja' espalhado categoria a categoria - mesma
    logica do arquivo real de Forecast (Detalhe_Despesas_Fitted Units,
    aba 'Resumo Custos'): valor_com_rateio = valor_proprio +
    (valor_da_Gerência NESSA MESMA categoria * % da unidade). Nao e' mais
    uma linha unica de rateio - cada categoria (Labour, Handling,
    Depreciation etc.) recebe o pedaco dela.

    So' o custo FIXO da Gerência entra no rateio (ver _apenas_fixo) - o
    Variável das unidades nunca e' afetado, porque a Gerência não tem
    custo Variável de verdade (confirmado pela usuária, 2026-08-25)."""
    ger = _apenas_fixo(totais_ativos.get(SIGLA_GERENCIA, {}), log, SIGLA_GERENCIA)
    resultado = {}
    for sigla in ORDEM_UNIDADES_ATIVAS:
        proprio = totais_ativos.get(sigla, {})
        pct = percentuais.get(sigla, 0)
        chaves = set(proprio.keys()) | set(ger.keys())
        resultado[sigla] = {
            chave: proprio.get(chave, 0) + ger.get(chave, 0) * pct for chave in chaves
        }
    return resultado


# ---------------------------------------------------------------------------
# Geracao do arquivo Excel
# ---------------------------------------------------------------------------

AZUL_ESCURO = "1F3864"
AZUL_CLARO_TOTAL = "D9E2F3"
CINZA_CLARO = "F2F2F2"
BRANCO = "FFFFFF"
FONTE_CABECALHO = Font(name="Calibri", size=11, bold=True, color=BRANCO)
FONTE_CATEGORIA = Font(name="Calibri", size=10, bold=True)
FONTE_ITEM = Font(name="Calibri", size=10, italic=True)
FONTE_TOTAL = Font(name="Calibri", size=10, bold=True)
FONTE_NOTA = Font(name="Calibri", size=9, italic=True, color="555555")
BORDA_FINA = Border(bottom=Side(style="thin", color="BFBFBF"))


def _escrever_quadro(ws, linha_inicio, col_inicio, titulo, colunas, dados, ordem_var, ordem_fixo, coluna_total=True):
    """Escreve um quadro (Variable Cost + Fixed Cost + Total Costs) a
    partir da linha/coluna dadas. `dados` = {coluna: {subcategoria: valor}}.
    Devolve a linha seguinte livre depois do quadro."""
    n_cols = len(colunas) + (1 if coluna_total else 0)

    # Cabecalho
    ws.merge_cells(start_row=linha_inicio, start_column=col_inicio, end_row=linha_inicio, end_column=col_inicio + n_cols)
    cel = ws.cell(row=linha_inicio, column=col_inicio, value=titulo)
    cel.font = Font(name="Calibri", size=13, bold=True)
    cel.alignment = Alignment(horizontal="center")

    linha_nota = linha_inicio + 1
    ws.cell(row=linha_nota, column=col_inicio, value="(+) gain   |   '000 BRL").font = FONTE_NOTA

    linha_cab = linha_inicio + 2
    ws.cell(row=linha_cab, column=col_inicio, value="")
    for i, colname in enumerate(colunas):
        c = ws.cell(row=linha_cab, column=col_inicio + 1 + i, value=colname)
        c.font = FONTE_CABECALHO
        c.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        c.alignment = Alignment(horizontal="center")
    if coluna_total:
        c = ws.cell(row=linha_cab, column=col_inicio + 1 + len(colunas), value="TOTAL")
        c.font = FONTE_CABECALHO
        c.fill = PatternFill("solid", fgColor=AZUL_ESCURO)
        c.alignment = Alignment(horizontal="center")

    linha = linha_cab + 1
    linhas_dados_totais = []  # (linha, e' linha de "Total Costs"?)

    def _linha_categoria(nome):
        nonlocal linha
        c = ws.cell(row=linha, column=col_inicio, value=nome)
        c.font = FONTE_CATEGORIA
        c.fill = PatternFill("solid", fgColor=CINZA_CLARO)
        primeira_linha_cat = linha
        linha += 1
        return primeira_linha_cat

    def _linha_item(nome, tipo_macro):
        # tipo_macro = "V" ou "F" - "Labour" (e outros nomes) existem nos
        # dois macros com valores DIFERENTES, por isso a chave em `dados` e'
        # sempre (tipo, nome), nunca so' o nome.
        nonlocal linha
        c = ws.cell(row=linha, column=col_inicio, value=nome)
        c.font = FONTE_ITEM
        for i, colname in enumerate(colunas):
            valor = dados.get(colname, {}).get((tipo_macro, nome), 0)
            # Guarda o valor CHEIO na célula (não arredondado) - só o
            # number_format exibe 1 casa decimal. Arredondar o VALOR (não só
            # a exibição) fazia o Check final (soma de quadro1 - soma de
            # quadro2) sobrar R$ 0,10 mil em vez de R$ 0,00 - cada célula
            # arredondada isoladamente antes de somar (~50 células) acumula
            # deriva de arredondamento, mesmo a matemática de verdade batendo
            # exato (achado real, testando contra a rede, 2026-08-26).
            cc = ws.cell(row=linha, column=col_inicio + 1 + i, value=valor)
            cc.font = FONTE_ITEM
            cc.number_format = "#,##0.0;(#,##0.0)"
        linha += 1

    # Variable Cost
    linha_var_categoria = _linha_categoria("Variable Cost")
    linha_var_inicio_itens = linha
    for item in ordem_var:
        _linha_item(item, "V")
    linha_var_fim_itens = linha - 1

    linha += 1  # linha em branco separando Variable Cost de Fixed Cost

    # Fixed Cost
    linha_fix_categoria = _linha_categoria("Fixed Cost")
    linha_fix_inicio_itens = linha
    for item in ordem_fixo:
        _linha_item(item, "F")
    linha_fix_fim_itens = linha - 1

    # Preenche a soma da linha "Variable Cost"/"Fixed Cost" (soma dos itens
    # dela) - cinza claro e negrito, igual o rotulo (pedido da usuaria).
    for i, colname in enumerate(colunas):
        col_letra = get_column_letter(col_inicio + 1 + i)
        cv = ws.cell(
            row=linha_var_categoria, column=col_inicio + 1 + i,
            value=f"=SUM({col_letra}{linha_var_inicio_itens}:{col_letra}{linha_var_fim_itens})",
        )
        cv.font = FONTE_CATEGORIA
        cv.fill = PatternFill("solid", fgColor=CINZA_CLARO)
        cv.number_format = "#,##0.0;(#,##0.0)"
        cf = ws.cell(
            row=linha_fix_categoria, column=col_inicio + 1 + i,
            value=f"=SUM({col_letra}{linha_fix_inicio_itens}:{col_letra}{linha_fix_fim_itens})",
        )
        cf.font = FONTE_CATEGORIA
        cf.fill = PatternFill("solid", fgColor=CINZA_CLARO)
        cf.number_format = "#,##0.0;(#,##0.0)"

    return {
        "linha_var_categoria": linha_var_categoria,
        "linha_fix_categoria": linha_fix_categoria,
        "linha_apos_itens": linha,
        "colunas": colunas,
        "col_inicio": col_inicio,
        "n_cols": n_cols,
        "coluna_total": coluna_total,
    }


def _finalizar_quadro_com_total(ws, info, linhas_extra_antes_do_total=None):
    """Adiciona (se pedido) linhas extras entre Fixed Cost e Total Costs
    (ex: 'Rateio Gerência'), depois a linha Total Costs (= Variable + Fixed
    + extras), e a coluna TOTAL (soma das unidades) em toda a área de dados."""
    linha = info["linha_apos_itens"]
    col_inicio = info["col_inicio"]
    colunas = info["colunas"]

    linhas_extras_num = []
    if linhas_extra_antes_do_total:
        for nome, valores in linhas_extra_antes_do_total:
            c = ws.cell(row=linha, column=col_inicio, value=nome)
            c.font = FONTE_ITEM
            for i, colname in enumerate(colunas):
                # Mesmo motivo da _linha_item: valor cheio na célula, só a
                # exibição arredonda - essa linha extra também entra na soma
                # de "Total Costs" (ver partes/linhas_extras_num abaixo).
                cc = ws.cell(row=linha, column=col_inicio + 1 + i, value=valores.get(colname, 0))
                cc.font = FONTE_ITEM
                cc.number_format = "#,##0.0;(#,##0.0)"
            linhas_extras_num.append(linha)
            linha += 1

    linha_total_costs = linha
    ws.cell(row=linha_total_costs, column=col_inicio, value="Total Costs").font = FONTE_TOTAL
    for i, colname in enumerate(colunas):
        col_letra = get_column_letter(col_inicio + 1 + i)
        partes = [f"{col_letra}{info['linha_var_categoria']}", f"{col_letra}{info['linha_fix_categoria']}"]
        partes += [f"{col_letra}{r}" for r in linhas_extras_num]
        cc = ws.cell(row=linha_total_costs, column=col_inicio + 1 + i, value="=" + "+".join(partes))
        cc.font = FONTE_TOTAL
        cc.number_format = "#,##0.0;(#,##0.0)"
        cc.border = Border(top=Side(style="thin", color="000000"))

    # Coluna TOTAL: soma das unidades, linha a linha, da primeira linha de
    # categoria ate a linha Total Costs.
    if info["coluna_total"]:
        col_total_letra = get_column_letter(col_inicio + 1 + len(colunas))
        primeira = info["linha_var_categoria"]
        for r in range(primeira, linha_total_costs + 1):
            if ws.cell(row=r, column=col_inicio).value is None:
                continue  # linha em branco (separador Variable/Fixed Cost)
            letras = [get_column_letter(col_inicio + 1 + i) for i in range(len(colunas))]
            formula = "=" + "+".join(f"{l}{r}" for l in letras)
            cc = ws.cell(row=r, column=col_inicio + 1 + len(colunas), value=formula)
            cc.number_format = "#,##0.0;(#,##0.0)"
            cc.fill = PatternFill("solid", fgColor=AZUL_CLARO_TOTAL)
            if r == linha_total_costs:
                cc.font = FONTE_TOTAL
                cc.border = Border(top=Side(style="thin", color="000000"))

    return linha_total_costs, linha + 1


def _escrever_linha_check(ws, linha, col_inicio, colunas, checks):
    """Escreve, logo abaixo do Total Costs do quadro "sem rateio", uma linha
    de check unidade por unidade: "✓" se o total bater com o valor bruto da
    Base Intermediária pra aquele Mini-Fábrica (diferença < R$ 0,05 mil,
    tolerância de arredondamento), ou um aviso com a diferença em reais se
    não bater - pedido explícito da usuária, 2026-08-26."""
    c = ws.cell(row=linha, column=col_inicio, value="Check (bate c/ Base Interm.)")
    c.font = FONTE_NOTA
    for i, colname in enumerate(colunas):
        diff = checks.get(colname, (0, 0, 0))[0]
        cc = ws.cell(row=linha, column=col_inicio + 1 + i)
        if abs(diff) < 0.05:
            cc.value = "✓"
            cc.font = Font(name="Calibri", size=10, bold=True, color="2E7D32")
        else:
            cc.value = f"⚠ R$ {diff * 1000:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            cc.font = Font(name="Calibri", size=9, bold=True, color="9C0006")
        cc.alignment = Alignment(horizontal="center")
    return linha + 1


def _escrever_aba_comentarios(wb, checks, fora_de_escopo, contas_nao_mapeadas):
    """Aba "Comentários" - reúne tudo que merece atenção da usuária num só
    lugar (pedido explícito, 2026-08-26), em vez de ficar só no log:
    resultado do check por unidade, linhas fora de escopo (Mini-Fábrica/
    Centro de Custo não reconhecido - antes descartadas em silêncio) e
    contas cuja Conta Geral não bateu com nenhuma subcategoria esperada."""
    ws = wb.create_sheet("Comentários")
    ws.sheet_view.showGridLines = False
    linha = 1
    ws.cell(row=linha, column=1, value="Comentários e verificações").font = Font(size=13, bold=True)
    linha += 2

    ws.cell(row=linha, column=1, value="Check por unidade (Total classificado vs. total bruto da Base Intermediária)").font = Font(bold=True)
    linha += 1
    cab = ["Unidade", "Total bruto (Base Interm.)", "Total classificado (quadro)", "Diferença (R$)", "Status"]
    for i, cnome in enumerate(cab):
        ws.cell(row=linha, column=1 + i, value=cnome).font = Font(bold=True)
    linha += 1
    for sigla in ORDEM_UNIDADES_ATIVAS + [SIGLA_GERENCIA]:
        diff, esperado, classificado = checks.get(sigla, (0, 0, 0))
        ws.cell(row=linha, column=1, value=sigla)
        c2 = ws.cell(row=linha, column=2, value=round(esperado, 1))
        c2.number_format = "#,##0.0;(#,##0.0)"
        c3 = ws.cell(row=linha, column=3, value=round(classificado, 1))
        c3.number_format = "#,##0.0;(#,##0.0)"
        c4 = ws.cell(row=linha, column=4, value=round(diff * 1000, 2))
        c4.number_format = "#,##0.00;(#,##0.00)"
        status = "✓ OK" if abs(diff) < 0.05 else "⚠ Verificar"
        cs = ws.cell(row=linha, column=5, value=status)
        if abs(diff) >= 0.05:
            cs.font = Font(bold=True, color="9C0006")
        linha += 1
    linha += 2

    if fora_de_escopo:
        ws.cell(row=linha, column=1, value="⚠ Linhas fora de escopo (Mini-Fábrica/Centro de Custo não reconhecido - não entraram no quadro)").font = Font(bold=True, color="9C0006")
        linha += 1
        cab2 = ["Conta", "Descrição", "Mini-Fábrica", "Centro de Custo", "Valor ('000 BRL)"]
        for i, cnome in enumerate(cab2):
            ws.cell(row=linha, column=1 + i, value=cnome).font = Font(bold=True)
        linha += 1
        for f in fora_de_escopo:
            ws.cell(row=linha, column=1, value=f["conta"])
            ws.cell(row=linha, column=2, value=f["descricao"])
            ws.cell(row=linha, column=3, value=f["mini_fabrica"])
            ws.cell(row=linha, column=4, value=f["centro_custo"])
            cv = ws.cell(row=linha, column=5, value=round(f["valor"], 1))
            cv.number_format = "#,##0.0;(#,##0.0)"
            linha += 1
        linha += 2
    else:
        ws.cell(row=linha, column=1, value="Nenhuma linha fora de escopo neste mês.").font = FONTE_NOTA
        linha += 2

    if contas_nao_mapeadas:
        ws.cell(row=linha, column=1, value="Contas com Conta Geral (AJ) não reconhecida - caíram em 'Other Variable/Fixed' por padrão").font = Font(bold=True, color="9C0006")
        linha += 1
        cab3 = ["Tipo (Var.)", "Conta Geral (AJ)", "Descrição"]
        for i, cnome in enumerate(cab3):
            ws.cell(row=linha, column=1 + i, value=cnome).font = Font(bold=True)
        linha += 1
        for tipo, conta_geral, desc in sorted(contas_nao_mapeadas, key=lambda x: (x[0], str(x[1]))):
            ws.cell(row=linha, column=1, value=tipo)
            ws.cell(row=linha, column=2, value=conta_geral)
            ws.cell(row=linha, column=3, value=desc)
            linha += 1
    else:
        ws.cell(row=linha, column=1, value="Nenhuma conta com Conta Geral não reconhecida neste mês.").font = FONTE_NOTA

    for col, largura in [(1, 14), (2, 40), (3, 16), (4, 18), (5, 16)]:
        ws.column_dimensions[get_column_letter(col)].width = largura


def gerar_arquivo_rateio_custos(mes: int, ano: int, ciclo: str, pasta_saida: Path, log=print) -> Path:
    caminho_base = localizar_base_intermediaria(mes, ano, ciclo)
    log(f"Lendo Base Intermediária: {caminho_base.name}...")
    totais_ativos, residuos_encerradas, contas_nao_mapeadas, raw_por_mini_fabrica, fora_de_escopo = (
        ler_e_classificar(caminho_base, mes, log)
    )
    checks = calcular_check_por_unidade(totais_ativos, raw_por_mini_fabrica, residuos_encerradas)
    for sigla, (diff, esperado, classificado) in checks.items():
        if abs(diff) >= 0.05:
            log(
                f"AVISO: check de {sigla} não bateu (diferença de {_fmt_moeda(diff)} mil) — "
                "ver aba 'Comentários' no arquivo gerado."
            )

    percentuais, vigente_desde = carregar_rateio_vigente(mes, ano)
    log(f"Rateio vigente desde {vigente_desde}: {percentuais}")

    rateio_por_unidade, gerencia_total, residuo_somado, residuo_sorocaba = calcular_rateio(
        totais_ativos, residuos_encerradas, percentuais, log
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Rateio de Custos"
    ws.sheet_view.showGridLines = False

    titulo_periodo = f"{MESES_ABREV[mes]} {ano} - {ciclo}"

    # --- Bloco 1: tabela do rateio vigente ---
    ws.cell(row=1, column=1, value=f"Rateio de Custos — {titulo_periodo}").font = Font(size=14, bold=True)
    ws.cell(row=2, column=1, value=f"Rateio Gerência vigente desde {vigente_desde}").font = FONTE_NOTA

    linha = 4
    ws.cell(row=linha, column=1, value="CM").font = FONTE_CABECALHO
    ws.cell(row=linha, column=1).fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    ws.cell(row=linha, column=2, value="Rateio Atual").font = FONTE_CABECALHO
    ws.cell(row=linha, column=2).fill = PatternFill("solid", fgColor=AZUL_ESCURO)
    linha += 1
    linha_pct_inicio = linha
    for sigla in ORDEM_UNIDADES_ATIVAS:
        ws.cell(row=linha, column=1, value=sigla)
        cc = ws.cell(row=linha, column=2, value=percentuais.get(sigla, 0))
        cc.number_format = "0.0%"
        linha += 1
    linha_pct_fim = linha - 1
    ws.cell(row=linha, column=1, value="TOTAL").font = FONTE_TOTAL
    cc = ws.cell(row=linha, column=2, value=f"=SUM(B{linha_pct_inicio}:B{linha_pct_fim})")
    cc.number_format = "0.0%"
    cc.font = FONTE_TOTAL

    # --- Avisos de resíduo (unidades encerradas) ---
    linha_aviso = linha + 3
    if residuos_encerradas:
        ws.cell(row=linha_aviso, column=1, value="⚠ Resíduo de unidade(s) encerrada(s) neste mês").font = Font(bold=True, color="9C0006")
        linha_aviso += 1
        cab = ["Unidade", "Conta", "Descrição", "Valor ('000 BRL)", "Entrou no rateio?"]
        for i, c in enumerate(cab):
            ws.cell(row=linha_aviso, column=1 + i, value=c).font = Font(bold=True)
        linha_aviso += 1
        for r in residuos_encerradas:
            entrou = "Não (Sorocaba — reclassificação não recorrente)" if r["unidade"].upper() == "SOROCABA" else "Sim (somado à Gerência)"
            ws.cell(row=linha_aviso, column=1, value=r["unidade"])
            ws.cell(row=linha_aviso, column=2, value=r["conta"])
            ws.cell(row=linha_aviso, column=3, value=r["descricao"])
            cc = ws.cell(row=linha_aviso, column=4, value=round(r["valor"], 1))
            cc.number_format = "#,##0.0;(#,##0.0)"
            ws.cell(row=linha_aviso, column=5, value=entrou)
            linha_aviso += 1
        linha_prox_bloco = linha_aviso + 2
    else:
        linha_prox_bloco = linha_aviso

    # --- Bloco 2: quadro sem rateio (SJP | IBI | GOI | RES | GER | TOTAL) ---
    colunas_sem_rateio = ORDEM_UNIDADES_ATIVAS + [SIGLA_GERENCIA]
    tem_nao_classificado = any(
        totais_ativos.get(c, {}).get("Não Classificado", 0) for c in colunas_sem_rateio
    )
    valores_nao_classificado = {
        c: totais_ativos.get(c, {}).get("Não Classificado", 0) for c in colunas_sem_rateio
    }
    # Nota: o resíduo de unidade encerrada (exceto Sorocaba) já foi somado
    # direto na categoria certa de totais_ativos[GER] (ler_e_classificar) -
    # não precisa de linha extra aqui, já aparece naturalmente dentro das
    # categorias da coluna GER.
    extras_quadro1 = [("Não Classificado", valores_nao_classificado)] if tem_nao_classificado else None

    info1 = _escrever_quadro(
        ws, linha_prox_bloco, 1,
        f"{titulo_periodo} — por unidade (sem rateio)",
        colunas_sem_rateio, totais_ativos, ORDEM_VARIAVEL, ORDEM_FIXO,
    )
    linha_total1, linha_prox_bloco = _finalizar_quadro_com_total(ws, info1, linhas_extra_antes_do_total=extras_quadro1)
    linha_prox_bloco = _escrever_linha_check(ws, linha_prox_bloco, info1["col_inicio"], colunas_sem_rateio, checks)
    linha_prox_bloco += 2

    # --- Bloco 3: quadro com rateio (SJP | IBI | GOI | RES | TOTAL) ---
    # O custo da Gerência já foi espalhado categoria a categoria (mesma
    # lógica do arquivo real de Forecast) - não é mais uma linha única.
    dados_com_rateio = calcular_dados_com_rateio(totais_ativos, percentuais, log)
    tem_nao_classificado_ativas = any(
        dados_com_rateio.get(c, {}).get("Não Classificado", 0) for c in ORDEM_UNIDADES_ATIVAS
    )
    extras_quadro2 = None
    if tem_nao_classificado_ativas:
        extras_quadro2 = [(
            "Não Classificado",
            {c: dados_com_rateio.get(c, {}).get("Não Classificado", 0) for c in ORDEM_UNIDADES_ATIVAS},
        )]

    info2 = _escrever_quadro(
        ws, linha_prox_bloco, 1,
        f"{titulo_periodo} — por unidade (com rateio)",
        ORDEM_UNIDADES_ATIVAS, dados_com_rateio, ORDEM_VARIAVEL, ORDEM_FIXO,
    )
    linha_total2, linha_prox_bloco = _finalizar_quadro_com_total(
        ws, info2, linhas_extra_antes_do_total=extras_quadro2
    )

    # --- Linha informativa "Rateio Gerência" (fora do Total Costs - o
    # rateio de verdade já está espalhado categoria a categoria acima; essa
    # linha só mostra, de forma resumida, quanto cada unidade recebeu no
    # total) - cinza claro, pedido da usuária. ---
    linha_prox_bloco += 1
    c = ws.cell(row=linha_prox_bloco, column=1, value="Rateio Gerência")
    c.font = FONTE_ITEM
    c.fill = PatternFill("solid", fgColor=CINZA_CLARO)
    for i, sigla in enumerate(ORDEM_UNIDADES_ATIVAS):
        cc = ws.cell(row=linha_prox_bloco, column=2 + i, value=round(rateio_por_unidade.get(sigla, 0), 1))
        cc.font = FONTE_ITEM
        cc.fill = PatternFill("solid", fgColor=CINZA_CLARO)
        cc.number_format = "#,##0.0;(#,##0.0)"
    col_letras = [get_column_letter(2 + i) for i in range(len(ORDEM_UNIDADES_ATIVAS))]
    cc = ws.cell(
        row=linha_prox_bloco, column=2 + len(ORDEM_UNIDADES_ATIVAS),
        value="=" + "+".join(f"{l}{linha_prox_bloco}" for l in col_letras),
    )
    cc.font = FONTE_ITEM
    cc.fill = PatternFill("solid", fgColor=CINZA_CLARO)
    cc.number_format = "#,##0.0;(#,##0.0)"
    linha_prox_bloco += 2

    # --- Check: TOTAL do quadro 1 (coluna TOTAL, linha Total Costs) deve
    # bater com o TOTAL do quadro 2 (coluna TOTAL, linha Total Costs) ---
    col_total1 = get_column_letter(info1["col_inicio"] + 1 + len(info1["colunas"]))
    col_total2 = get_column_letter(info2["col_inicio"] + 1 + len(info2["colunas"]))
    ws.cell(row=linha_prox_bloco + 1, column=1, value="Check (deve ser 0,00):").font = Font(bold=True)
    cc = ws.cell(
        row=linha_prox_bloco + 1, column=2,
        value=f"={col_total1}{linha_total1}-{col_total2}{linha_total2}",
    )
    cc.number_format = "#,##0.00;(#,##0.00)"
    cc.font = Font(bold=True)

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 16 if col > 1 else 22

    _escrever_aba_comentarios(wb, checks, fora_de_escopo, contas_nao_mapeadas)

    pasta_saida.mkdir(parents=True, exist_ok=True)
    nome_arquivo = nome_com_versao(pasta_saida, f"Rateio de Custos Fitted Units {MESES_INGLES[mes]} {ciclo} {ano}.xlsx")
    caminho_saida = pasta_saida / nome_arquivo
    wb.save(caminho_saida)
    log(f"Arquivo de Rateio de Custos gerado: {caminho_saida}")
    return caminho_saida


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Gera o arquivo de Rateio de Custos (Passo 5)")
    parser.add_argument("--mes", type=int, required=True)
    parser.add_argument("--ano", type=int, required=True)
    parser.add_argument("--ciclo", choices=["Actual", "Flash"], required=True)
    parser.add_argument("--pasta-saida", type=Path, required=True, help="Pasta de saída (teste local ou rede)")
    args = parser.parse_args()

    gerar_arquivo_rateio_custos(args.mes, args.ano, args.ciclo, args.pasta_saida)
