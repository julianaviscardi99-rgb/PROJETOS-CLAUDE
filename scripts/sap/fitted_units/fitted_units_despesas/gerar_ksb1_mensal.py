#!/usr/bin/env python3
"""
Automatiza o passo 3 do processo recorrente de Fitted Units Despesas: monta
o arquivo mensal "KSB1 <Mes> <Ciclo> <Ano>.xlsx" a partir do arquivo Actual
do mes anterior, replicando o processo manual real (BASE_KSB1 + Pivot nativo
- ver memory/DECISOES.md, reversao de arquitetura de 2026-08-11):

1. Copia o KSB1 do Actual do mes anterior (ja acumulado) -> nova copia
   versionada (nunca sobrescreve o arquivo original).
2. Decide Gestoriais vs Sem Agrupamento (mesma regra do check_agrupamentos)
   e cola as linhas novas do extrato bruto do mes no fim da aba BASE_KSB1
   (colunas A-R, mapeamento 1:1 confirmado com o extrato).
3. Arrasta (AutoFill, equivalente a "arrastar a formula") as colunas S:AI
   da ultima linha existente para as linhas novas.
4. Da refresh nas Pivot Tables nativas (Pivot_Inter., Pivot_Detalhes) via
   automacao COM do Excel.
5. Salva a copia.

Links externos do BASE_KSB1 (base de contas: abas Contas / Centros) SAO
resolvidos ao vivo a cada geracao desde 2026-09-04: o arquivo de origem e'
aberto (somente leitura) na mesma instancia do Excel enquanto o arquivo do mes
e' montado e salvo - ver abrir_fontes_dos_links().
Antes disso eram deliberadamente ignorados (UpdateLinks=0), pra nao misturar
"mudanca na base de contas" com "erro na automacao" ao validar contra um mes
ja fechado manualmente; mas em producao isso fazia conta nova cadastrada no
de-para NUNCA resolver (o arquivo do mes herda o cache do mes anterior), o que
custou R$ 79.787,48 de custo inflado no fechamento de Agosto/2026 Actual.

Depende de Excel instalado (pywin32) - abre uma instancia oculta e isolada
(DispatchEx), nao interfere com o Excel que a usuaria tiver aberto.
"""
import re
import shutil
import sys
import zipfile
from pathlib import Path

import pywintypes
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "_shared"))
from ksb1_core import (  # noqa: E402
    BU,
    MESES_PASTA,
    REDE_BASE,
    abrir_excel_isolado,
    com_retry,
    encontrar_arquivo_mais_recente,
    localizar_extracao_ksb1,
    nome_com_versao,
    resolver_pasta_ciclo,
)

sys.path.insert(0, str(Path(__file__).resolve().parent))
from check_agrupamentos_ksb1 import eh_conta_ignorada, linha_e_subtotal  # noqa: E402

MESES_INGLES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}

N_COLS_BRUTO = 18  # colunas A-R do BASE_KSB1, 1:1 com o extrato bruto da KSB1
COL_CONTA = 4
COL_VALOR = 17
COL_MES_BASE = 19       # S ("Mês")
COL_GESTORIAL_BASE = 20  # T ("Gestorial")

ABA_PIVOT_INTER = "Pivot_Inter."

# Coluna AH ("Centro de Montagem(2)"): resolve a MF (coluna Z) na tabela
# de-para da base de contas externa (Base_Contas_Contabeis_Fitted_22.xlsx,
# aba Centros, colunas K:L). A formula herdada do arquivo do mes anterior
# vinha com o range TRAVADO em $K$2:$L$9 - quando uma unidade nova e'
# cadastrada abaixo da linha 9 da tabela, ela fica FORA do range e toda
# linha dela resolve em #N/A, contaminando o Pivot_Inter. e, na sequencia,
# a Base Intermediaria (caso real: Resende / MF 0483, cadastrada na linha
# K10, quebrou o fechamento de Agosto/2026 - 2026-09-01).
COL_CENTRO_MONTAGEM_2 = 34  # AH
ULTIMA_LINHA_DEPARA_MF = 100

XL_UP = -4162
XL_FILL_DEFAULT = 0
XL_CALCULATION_MANUAL = -4135
XL_CALCULATION_AUTOMATIC = -4105
XL_CELL_TYPE_FORMULAS = -4123
XL_ERRORS = 16
XL_LINK_TYPE_EXCEL_LINKS = 1


def _abrev(mes: int) -> str:
    return MESES_PASTA[mes].split(" - ")[1]


def decidir_fonte_e_ler_linhas(mes: int, ano: int, ciclo: str, log=print):
    """Decide Gestoriais vs Sem Agrupamento (mesma regra do check_agrupamentos_ksb1)
    e devolve as linhas de detalhe completas (18 colunas, A-R do BASE_KSB1),
    ja filtradas (sem subtotal, sem conta em branco, excluindo contas do
    Check 1 se for usado o Sem Agrupamento). As extracoes brutas usadas sao
    sempre do Ciclo pedido (ver ksb1_core.encontrar_arquivo_ksb1) - nao mais
    "a mais recente por data de modificacao", que podia pegar por engano a
    extracao de outro Ciclo do mesmo mes."""
    pasta_mes = REDE_BASE / str(ano) / "00.Extração Base KSB1" / MESES_PASTA[mes]
    arquivo_gest = localizar_extracao_ksb1(pasta_mes, BU["nome"], mes, ano, "Gestoriais", ciclo)
    arquivo_sem = localizar_extracao_ksb1(pasta_mes, BU["nome"], mes, ano, "Sem Agrupamento", ciclo)

    def linhas_completas(caminho):
        wb = load_workbook(caminho, data_only=True)
        ws = wb.active
        linhas = []
        for r in range(2, ws.max_row + 1):
            conta = ws.cell(row=r, column=COL_CONTA).value
            if conta in (None, ""):
                continue
            if linha_e_subtotal(ws, r):
                continue
            linha = [ws.cell(row=r, column=c).value for c in range(1, N_COLS_BRUTO + 1)]
            valor = float(ws.cell(row=r, column=COL_VALOR).value or 0)
            linhas.append((str(conta).strip(), valor, linha))
        return linhas

    linhas_gest = linhas_completas(arquivo_gest)
    linhas_sem = linhas_completas(arquivo_sem)

    total_gest = sum(v for _, v, _ in linhas_gest)
    linhas_sem_filtradas = [(c, v, l) for c, v, l in linhas_sem if not eh_conta_ignorada(c, mes, ano)]
    total_sem_filtrado = sum(v for _, v, _ in linhas_sem_filtradas)

    if abs(total_gest - total_sem_filtrado) < 0.01:
        log(f"Gestoriais bate com Sem Agrupamento (total={total_gest:,.2f}) — usando Gestoriais ({arquivo_gest.name}).")
        return [l for _, _, l in linhas_gest], "Gestoriais", arquivo_gest
    else:
        log(
            f"Gestoriais ({total_gest:,.2f}) NÃO bate com Sem Agrupamento filtrado "
            f"({total_sem_filtrado:,.2f}) — usando Sem Agrupamento (excluindo contas do Check 1), "
            f"conforme regra confirmada em 2026-08-11 ({arquivo_sem.name})."
        )
        return [l for _, _, l in linhas_sem_filtradas], "Sem Agrupamento", arquivo_sem


def localizar_ksb1_actual_anterior(mes: int, ano: int) -> Path:
    mes_ant, ano_ant = (mes - 1, ano) if mes > 1 else (12, ano - 1)
    pasta_mes = REDE_BASE / str(ano_ant) / MESES_PASTA[mes_ant]
    pasta = resolver_pasta_ciclo(pasta_mes, mes_ant, "Actual")
    nome = f"KSB1 {MESES_INGLES[mes_ant]} Actual {ano_ant}.xlsx"
    caminho = encontrar_arquivo_mais_recente(pasta, nome)
    if caminho is None:
        raise FileNotFoundError(f"Não encontrei o KSB1 Actual do mês anterior: {pasta / nome}")
    return caminho


def copiar_para_teste(caminho_origem: Path, pasta_destino: Path, nome_base: str, log=print) -> Path:
    pasta_destino.mkdir(parents=True, exist_ok=True)
    nome_saida = nome_com_versao(pasta_destino, nome_base)
    caminho_saida = pasta_destino / nome_saida
    log(f"Copiando {caminho_origem.name} -> {caminho_saida} ...")
    shutil.copy2(caminho_origem, caminho_saida)
    return caminho_saida


def remover_flag_somente_leitura_recomendada(caminho: Path, log=print):
    """O KSB1 Actual/Flash tem a flag interna 'Somente leitura recomendada'
    (<fileSharing readOnlyRecommended="1"/> em xl/workbook.xml) gravada no
    arquivo. Com DisplayAlerts=False, o Excel abre silenciosamente em modo
    leitura mesmo com IgnoreReadOnlyRecommended=True no Open() (o parâmetro
    não se mostrou confiável via COM), e o Save() vira um no-op sem erro.
    Como esta função só roda na NOSSA cópia de teste (nunca no arquivo
    original), removemos a flag direto do XML antes de abrir no Excel."""
    with zipfile.ZipFile(caminho, "r") as zin:
        itens = {n: zin.read(n) for n in zin.namelist()}
        infos = {n: zin.getinfo(n) for n in zin.namelist()}

    wb_xml = itens["xl/workbook.xml"].decode("utf-8")
    nova_xml, n_removidas = re.subn(r"<fileSharing[^/]*/>", "", wb_xml)
    if n_removidas == 0:
        log("Arquivo não tinha a flag 'Somente leitura recomendada' — nada a remover.")
        return
    itens["xl/workbook.xml"] = nova_xml.encode("utf-8")

    with zipfile.ZipFile(caminho, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, data in itens.items():
            zout.writestr(infos[n], data)
    log(f"Removida a flag 'Somente leitura recomendada' da cópia de teste ({n_removidas} ocorrência(s)).")


def normalizar_formula_centro_montagem(ws, last_row: int, log) -> bool:
    """Amplia o range travado do de-para de MF na fórmula da coluna AH
    ('Centro de Montagem(2)') — ver comentário de COL_CENTRO_MONTAGEM_2.
    Reescreve a coluna inteira com UMA atribuição (o Excel ajusta sozinho a
    referência relativa de cada linha), então é barato mesmo com dezenas de
    milhares de linhas. Devolve True se mudou alguma coisa."""
    formula = ws.Cells(2, COL_CENTRO_MONTAGEM_2).Formula
    nova = re.sub(r"(\$K\$2:\$L\$)\d+", rf"\g<1>{ULTIMA_LINHA_DEPARA_MF}", formula)
    if nova == formula:
        return False
    ws.Range(
        ws.Cells(2, COL_CENTRO_MONTAGEM_2), ws.Cells(last_row, COL_CENTRO_MONTAGEM_2)
    ).Formula = nova
    log(
        f"Fórmula da coluna AH (Centro de Montagem) ampliada até a linha {ULTIMA_LINHA_DEPARA_MF} "
        "do de-para de MF — unidade nova cadastrada no fim da tabela não vira mais #N/A."
    )
    return True


def abrir_fontes_dos_links(excel, wb, log=print) -> list:
    """Abre (SOMENTE LEITURA) os arquivos de origem dos links externos do arquivo do
    mes - na pratica a base de contas (Base_Contas_Contabeis_Fitted_22.xlsx, abas
    'Contas' e 'Centros'). Devolve a lista de workbooks abertos, pra quem chama
    fechar depois de salvar.

    Por que isso e' necessario (bug real, 2026-09-04): o arquivo do mes e' uma COPIA
    do Actual do mes anterior, entao ele carrega o CACHE do link externo daquele mes.
    Sem isso, uma conta cadastrada no de-para DEPOIS daquele mes nunca resolve - a
    coluna T (Gestorial) fica #N/A e a Pivot_Inter. descarta a linha em silencio (o
    item '#N/A' esta desmarcado no filtro do campo). Foi assim que a conta M240600000
    (repasse de Ibirite, -R$ 79.787,48) sumiu do fechamento de Agosto/2026 Actual,
    deixando o custo R$ 79 mil mais alto que a realidade.

    Por que ABRIR a fonte em vez de chamar UpdateLink (ideia da usuaria, 2026-09-04,
    testada e confirmada): com o arquivo de origem aberto na mesma instancia do Excel,
    a formula resolve AO VIVO contra ele, sem depender do cache nem de o UpdateLink
    ter funcionado - UpdateLink e' uma chamada que pode retornar sem erro e nao ter
    atualizado nada. Fica UpdateLink so' como plano B, se a fonte nao puder ser aberta.

    IMPORTANTE pra quem chama: so' feche estes workbooks DEPOIS de salvar o arquivo do
    mes - fechada a fonte antes do Save, o Excel volta a usar o valor em cache.

    Falha aqui nao derruba a geracao: loga aviso e segue com o cache (o BASE_KSB1 ja
    teve link quebrado pra RHFitted, confirmado como lixo em 2026-08-11). Se o de-para
    nao tiver resolvido, conferir_pivot_contra_base() pega a divergencia no fim."""
    fontes = wb.LinkSources(XL_LINK_TYPE_EXCEL_LINKS)
    if not fontes:
        log("Nenhum link externo pra resolver.")
        return []

    abertos = []
    for fonte in fontes:
        caminho = Path(str(fonte))
        if not caminho.exists():
            log(f"AVISO: fonte do link não encontrada, seguindo com o cache: {caminho.name}")
            continue
        try:
            abertos.append(
                com_retry(
                    excel.Workbooks.Open,
                    str(caminho), UpdateLinks=0, ReadOnly=True,
                    log=log,
                )
            )
            log(f"Fonte do link aberta (somente leitura): {caminho.name}")
        except pywintypes.com_error as e:
            log(f"AVISO: não consegui abrir a fonte {caminho.name} ({e}) — tentando UpdateLink.")
            try:
                com_retry(wb.UpdateLink, Name=fonte, Type=XL_LINK_TYPE_EXCEL_LINKS, log=log)
                log(f"Link atualizado via UpdateLink: {caminho.name}")
            except pywintypes.com_error as e2:
                log(
                    f"AVISO: UpdateLink também falhou para {caminho.name} ({e2}) — "
                    "as fórmulas que dependem dele seguem com o valor em cache."
                )
    return abertos


def fechar_fontes_dos_links(abertos: list, log=print):
    """Fecha os workbooks de origem abertos por abrir_fontes_dos_links().
    Chamar SEMPRE depois de salvar o arquivo do mes."""
    for fonte in abertos:
        try:
            com_retry(fonte.Close, SaveChanges=False, log=log)
        except pywintypes.com_error:
            log("AVISO: não consegui fechar um dos arquivos de origem dos links.")


def _contas_com_gestorial_em_erro(ws, last_row: int, mes: int) -> list:
    """Lista as contas contabeis do mes cujo Gestorial (coluna T) resolveu em erro
    (#N/A tipicamente: conta que nao existe no de-para 'Contas' da base de contas).
    Usa SpecialCells pra pegar so' as celulas em erro - nao varre as 60+ mil linhas."""
    try:
        celulas = ws.Range(
            ws.Cells(2, COL_GESTORIAL_BASE), ws.Cells(last_row, COL_GESTORIAL_BASE)
        ).SpecialCells(XL_CELL_TYPE_FORMULAS, XL_ERRORS)
    except pywintypes.com_error:
        return []  # SpecialCells estoura quando nao ha nenhuma celula em erro

    if celulas.Count > 500:  # algo muito errado - nao vale varrer celula a celula
        return []

    achados = {}
    for celula in celulas:
        linha = celula.Row
        if ws.Cells(linha, COL_MES_BASE).Value != mes:
            continue
        conta = ws.Cells(linha, COL_CONTA).Value
        valor = ws.Cells(linha, COL_VALOR).Value or 0
        atual = achados.setdefault(conta, [0.0, 0])
        atual[0] += valor
        atual[1] += 1
    return sorted(achados.items(), key=lambda item: item[1][0])


def conferir_pivot_contra_base(excel, wb, ws, last_row: int, mes: int, log=print):
    """Confere se o Grand Total do mes na Pivot_Inter. bate com a soma direta do
    BASE_KSB1 daquele mes.

    Motivo (bug real, 2026-09-04, Agosto/2026 Actual): o filtro do campo "Gestorial"
    da Pivot_Inter. tem os itens `#N/A` e `(vazio)` DESMARCADOS. Qualquer linha cujo
    Gestorial nao resolva (conta nova fora do de-para) some do Grand Total sem erro
    nenhum - naquele caso, 2 linhas de credito (-79.787,48, repasse de Ibirite,
    conta M240600000) sumiram e a Base Intermediaria herdou custo INFLADO em R$ 79 mil.
    E' o mesmo padrao do bug das provisoes de 2026-09-02 (item `(blank)` do campo
    "Var." desmarcado). Esta conferencia pega essa classe inteira de problema de uma
    vez, qualquer que seja o campo/item escondido no filtro.

    Nao aborta antes de salvar: os dados da BASE_KSB1 estao corretos e a colagem
    custa 10+ minutos - o arquivo e' preservado e o erro sobe depois, pra usuaria
    resolver o de-para antes de seguir pra Finalizacao da Base Intermediaria."""
    total_base = com_retry(
        excel.WorksheetFunction.SumIf,
        ws.Columns(COL_MES_BASE), mes, ws.Columns(COL_VALOR),
        log=log,
    )

    wsp = wb.Worksheets(ABA_PIVOT_INTER)
    pt = wsp.PivotTables(1)
    intervalo = pt.TableRange1
    linha_cabecalho = intervalo.Row + 1
    linha_grand_total = intervalo.Row + intervalo.Rows.Count - 1

    col_mes = None
    for i in range(intervalo.Columns.Count):
        coluna = intervalo.Column + i
        if wsp.Cells(linha_cabecalho, coluna).Value == mes:
            col_mes = coluna
    if col_mes is None:
        log(
            f"AVISO: não achei a coluna do mês {mes} na {ABA_PIVOT_INTER} — "
            "não foi possível conferir a Pivot contra o BASE_KSB1."
        )
        return

    total_pivot = wsp.Cells(linha_grand_total, col_mes).Value or 0.0
    diferenca = total_pivot - total_base
    if abs(diferenca) <= 0.01:
        log(f"Conferência OK: Pivot_Inter. e BASE_KSB1 batem no mês {mes} ({total_base:,.2f}).")
        return

    contas = _contas_com_gestorial_em_erro(ws, last_row, mes)
    if contas:
        detalhe = "; ".join(
            f"{conta} ({valor:,.2f} em {n} linha(s))" for conta, (valor, n) in contas
        )
        pista = (
            f"\nContas do mês com Gestorial em erro (#N/A) — provavelmente não cadastradas na aba "
            f"'Contas' da Base_Contas_Contábeis_Fitted: {detalhe}."
            f"\nCadastre a(s) conta(s) no de-para e gere o arquivo de novo. Atenção: o BASE_KSB1 abre "
            f"com os links externos NÃO atualizados, então também é preciso atualizar os links "
            f"(Dados > Editar Links > Atualizar valores) pra fórmula reresolver."
        )
    else:
        pista = (
            "\nNão achei célula em erro na coluna Gestorial — verifique os FILTROS dos campos da "
            "Pivot_Inter. (itens desmarcados, como '#N/A' ou '(vazio)', somem do Grand Total sem avisar)."
        )

    raise RuntimeError(
        f"O arquivo FOI SALVO e os dados da BASE_KSB1 estão corretos, mas a Pivot_Inter. NÃO bate "
        f"com eles no mês {mes}:\n"
        f"  BASE_KSB1 (soma real): {total_base:,.2f}\n"
        f"  Pivot_Inter. (Grand Total): {total_pivot:,.2f}\n"
        f"  Diferença: {diferenca:,.2f}\n"
        f"A Pivot está deixando linha(s) de fora — NÃO siga para a 'Finalização da Base Intermediária' "
        f"antes de resolver, senão a Base Intermediária herda o valor errado.{pista}"
    )


def colar_linhas_e_atualizar_pivots(caminho_copia: Path, linhas_novas: list, mes: int, log=print, pid_callback=None):
    excel = abrir_excel_isolado(log, pid_callback)
    fontes_abertas = []
    try:
        # IgnoreReadOnlyRecommended=True: o BASE_KSB1 tem a flag interna
        # "Somente leitura recomendada" (fileSharing readOnlyRecommended="1")
        # - sem isso, com DisplayAlerts=False, o Excel abre o arquivo em modo
        # leitura silenciosamente e o Save() vira um no-op sem erro nenhum.
        wb = com_retry(
            excel.Workbooks.Open,
            str(caminho_copia), UpdateLinks=0, ReadOnly=False, IgnoreReadOnlyRecommended=True,
            log=log,
        )
        if wb.ReadOnly:
            raise RuntimeError(
                "O arquivo abriu em modo somente leitura mesmo com IgnoreReadOnlyRecommended=True "
                "— provavelmente já está aberto/travado por outro processo do Excel."
            )
        ws = wb.Worksheets("BASE_KSB1")
        last_row = ws.Cells(ws.Rows.Count, 1).End(XL_UP).Row
        n = len(linhas_novas)
        log(f"BASE_KSB1: última linha existente = {last_row}. Colando {n} linha(s) nova(s) (linhas {last_row + 1}-{last_row + n})...")

        # Calculo manual durante a colagem/AutoFill: com calculo automatico
        # (padrao do Excel), CADA uma das milhares de escritas linha-a-linha
        # abaixo dispara recalculo do arquivo inteiro (formulas + Pivot) -
        # e' o principal motivo da operacao levar 10+ minutos num mes com
        # muitas linhas (achado real, 2026-09-01, usuaria reportou "mais de
        # 12 minutos"). Nao afeta a protecao contra o bug de corrupcao
        # #N/A (que e' sobre GRANULARIDADE da escrita, nao sobre o modo de
        # calculo) - o recalculo completo continua acontecendo, so' que UMA
        # vez so' (CalculateFullRebuild, abaixo) em vez de milhares de vezes.
        # Restaurado pra automatico antes de salvar, pra nao mudar o modo de
        # calculo padrao do arquivo pra quem abrir depois.
        com_retry(setattr, excel, "Calculation", XL_CALCULATION_MANUAL, log=log)
        com_retry(setattr, excel, "ScreenUpdating", False, log=log)

        # Com o calculo ja em manual: abrir as fontes agora nao dispara recalculo
        # (o CalculateFullRebuild mais abaixo resolve tudo de uma vez). As fontes
        # ficam abertas ate DEPOIS do Save - ver abrir_fontes_dos_links().
        log("Abrindo as fontes dos links externos (base de contas)...")
        fontes_abertas = abrir_fontes_dos_links(excel, wb, log)

        # Colar linha por linha, NUNCA o bloco inteiro de uma vez: escrever um
        # array grande via Range.Value em uma unica chamada COM pode corromper
        # aleatoriamente algumas celulas em erro #N/A (bug de marshalling do
        # pywin32/COM, sem relacao com o conteudo - achado e confirmado
        # testando gerar_base_intermediaria.py ao vivo em 2026-08-21: 166
        # erros colando tudo de uma vez, 25 em blocos de 50, ZERO linha por
        # linha). Nenhum arquivo de producao foi afetado ate agora (confirmado
        # que este script nunca rodou de verdade contra a rede), mas a mesma
        # proteção foi aplicada aqui preventivamente.
        for i, linha in enumerate(linhas_novas):
            r = last_row + 1 + i
            ws.Range(ws.Cells(r, 1), ws.Cells(r, N_COLS_BRUTO)).Value = [linha]

        last_row_pos_escrita = ws.Cells(ws.Rows.Count, 1).End(XL_UP).Row
        if last_row_pos_escrita != last_row + n:
            raise RuntimeError(
                f"Depois de colar, a última linha em memória é {last_row_pos_escrita}, "
                f"esperava {last_row + n} — a escrita não aconteceu como esperado."
            )

        log("Verificando se alguma célula foi gravada como erro (#N/A) durante a colagem...")
        conferencia = ws.Range(ws.Cells(last_row + 1, 1), ws.Cells(last_row + n, N_COLS_BRUTO)).Value
        celulas_com_erro = sum(
            1 for linha in conferencia for v in linha if isinstance(v, int) and v < 0
        )
        if celulas_com_erro:
            raise RuntimeError(
                f"{celulas_com_erro} célula(s) gravada(s) como erro (#N/A) mesmo colando linha por linha — "
                "abortando sem salvar."
            )

        log("Arrastando fórmulas das colunas S:AI para as linhas novas (AutoFill)...")
        origem_formula = ws.Range(ws.Cells(last_row, 19), ws.Cells(last_row, 35))
        destino_formula = ws.Range(ws.Cells(last_row, 19), ws.Cells(last_row + n, 35))
        com_retry(origem_formula.AutoFill, destino_formula, XL_FILL_DEFAULT, log=log)

        com_retry(normalizar_formula_centro_montagem, ws, last_row + n, log, log=log)

        log("Recalculando a planilha inteira...")
        com_retry(excel.CalculateFullRebuild, log=log)

        # Restaura o modo padrao (automatico) antes de salvar/mexer nas
        # Pivots - o CalculateFullRebuild acima ja forcou o recalculo
        # completo uma vez, entao nao perde nada; so' evita salvar o arquivo
        # com o app deixado em modo manual.
        com_retry(setattr, excel, "Calculation", XL_CALCULATION_AUTOMATIC, log=log)
        com_retry(setattr, excel, "ScreenUpdating", True, log=log)

        log("Atualizando as Pivot Tables (Pivot_Inter., Pivot_Detalhes)...")
        com_retry(wb.RefreshAll, log=log)
        com_retry(excel.CalculateUntilAsyncQueriesDone, log=log)

        log("Salvando...")
        com_retry(wb.Save, log=log)
        if not wb.Saved:
            raise RuntimeError("wb.Save() retornou mas wb.Saved ainda é False — o arquivo pode não ter sido gravado.")

        # Conferencia DEPOIS de salvar, de proposito: se a Pivot nao bater, o
        # arquivo (que custou 10+ minutos de colagem) e' preservado e o erro
        # sobe pra usuaria resolver o de-para antes do proximo passo.
        log("Conferindo se a Pivot_Inter. bate com o BASE_KSB1...")
        conferir_pivot_contra_base(excel, wb, ws, last_row + n, mes, log)

        com_retry(wb.Close, SaveChanges=False, log=log)
        log("Concluído.")
    finally:
        # Fechar as fontes SO' aqui, depois do Save la' em cima: com a fonte
        # fechada antes de salvar, o Excel volta a gravar o valor em cache.
        fechar_fontes_dos_links(fontes_abertas, log)

        # Retentativa tambem no Quit - se o Excel "ocupado" derrubar essa
        # chamada, ela e' engolida sem retentar, deixa o processo orfao
        # rodando pra sempre (achado ao vivo, 2026-09-01: sobrou um
        # EXCEL.EXE ocioso depois de uma rodada bem-sucedida deste script).
        try:
            com_retry(excel.Quit, log=log)
        except pywintypes.com_error:
            log("AVISO: não consegui confirmar o encerramento do Excel (pode ter ficado um processo órfão).")


def gerar_ksb1_mensal(
    mes: int, ano: int, ciclo: str, pasta_saida: Path, sufixo_nome: str = "", log=print, pid_callback=None
) -> Path:
    linhas_novas, fonte, arquivo_fonte = decidir_fonte_e_ler_linhas(mes, ano, ciclo, log)

    caminho_origem = localizar_ksb1_actual_anterior(mes, ano)
    log(f"Partindo do Actual do mês anterior: {caminho_origem.name}")

    nome_base = f"KSB1 {MESES_INGLES[mes]} {ciclo} {ano}{sufixo_nome}.xlsx"
    caminho_copia = copiar_para_teste(caminho_origem, pasta_saida, nome_base, log)
    remover_flag_somente_leitura_recomendada(caminho_copia, log)

    colar_linhas_e_atualizar_pivots(caminho_copia, linhas_novas, mes, log, pid_callback)

    log(f"\nArquivo gerado: {caminho_copia}")
    log(f"Fonte usada: {fonte} ({arquivo_fonte.name}), {len(linhas_novas)} linha(s) colada(s).")
    return caminho_copia


if __name__ == "__main__":
    import sys as _sys

    if len(_sys.argv) != 4:
        print("Uso: python gerar_ksb1_mensal.py <mes> <ano> <Actual|Flash>")
        _sys.exit(1)
    _mes, _ano, _ciclo = int(_sys.argv[1]), int(_sys.argv[2]), _sys.argv[3]
    _pasta_saida = Path(__file__).resolve().parent.parent.parent.parent.parent / "data" / "processed" / "fitted_units_despesas" / "base_ksb1_teste"
    gerar_ksb1_mensal(_mes, _ano, _ciclo, _pasta_saida, sufixo_nome=" - TESTE VALIDAÇÃO")
