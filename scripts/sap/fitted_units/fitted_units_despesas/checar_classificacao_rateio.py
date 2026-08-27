#!/usr/bin/env python3
"""
Ferramenta de auditoria pro Passo 5 (Rateio de Custos) - NAO faz parte do
processo recorrente mensal, e' pra rodar sob demanda quando quiser
reconferir a classificacao Variavel/Fixo por conta.

Desenhada com a Juliana em 2026-08-26, depois dela ter dito "ainda estou
insegura e com medo de um dia classificarmos errado". Ideia dela: em vez
de investigar exceção por exceção quando aparece um problema, extrair a
estrutura COMPLETA do arquivo antigo "_Abertura custos Fitted Units <Mes>
<Ciclo> <Ano>.xlsx" (cada categoria/"voz" com todas as contas gestoriais
que pertencem a ela) e comparar, conta por conta, contra o que a
classificacao atual (AA/AJ da Base Intermediaria + excecoes em
gerar_rateio_custos.py) produziria hoje - vira uma "memoria de calculo"
completa, nao so' os casos que ja' causaram problema.

Uso (linha de comando):
    python checar_classificacao_rateio.py --arquivo-antigo "<caminho>.xlsx" --meses 1 2 3 4 5 6 7 --ano 2026

So' LE arquivos (win32com ReadOnly + openpyxl read_only) - nunca escreve
em nada.
"""
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import win32com.client

sys.path.insert(0, str(Path(__file__).resolve().parent))
from gerar_rateio_custos import (  # noqa: E402
    CONTAS_FORCADAS_DEPRECIATION,
    CONTAS_FORCADAS_HANDLING_VARIAVEL,
    CONTAS_FORCADAS_TRANSPORTATION_VARIAVEL,
    _resolver_subcategoria,
    localizar_base_intermediaria,
)

CATEGORIAS_VAR = {"Labour", "Handling", "Direct Materials", "Transportation", "Other Variable"}
CATEGORIAS_FIXO = {"Labour", "Depreciation", "IFRS16", "Rents", "Condominio", "Other Fixed"}
ABAS_UNIDADE_ARQUIVO_ANTIGO = ["São J. dos Pinhais", "Ibirité", "Goiana"]


def extrair_estrutura_arquivo_antigo(caminho_abertura_custos: Path, log=print) -> dict:
    """Le o arquivo '_Abertura custos...' e devolve {(conta, tipo):
    categoria} - a categoria/'voz' declarada na PRÓPRIA estrutura do
    arquivo (não inferida por valor). Só devolve combinações em que as 3
    abas de unidade concordam entre si (avisa no log se alguma divergir -
    nunca deveria acontecer, é o mesmo template pras 3)."""
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.AskToUpdateLinks = False
    mapa_por_unidade = {}
    try:
        wb = excel.Workbooks.Open(str(Path(caminho_abertura_custos).resolve()), ReadOnly=True)
        for aba in ABAS_UNIDADE_ARQUIVO_ANTIGO:
            ws = wb.Worksheets(aba)
            macro_atual = None
            categoria_atual = None
            mapa = {}
            for r in range(53, 208):
                nome = ws.Cells(r, 6).Value
                if nome is None:
                    continue
                nome = str(nome).strip()
                if nome == "Variable Cost":
                    macro_atual = "V"
                    continue
                if nome == "Fixed Cost":
                    macro_atual = "F"
                    continue
                if nome == "Total Costs":
                    break
                categorias_validas = CATEGORIAS_VAR if macro_atual == "V" else CATEGORIAS_FIXO
                if nome in categorias_validas:
                    categoria_atual = nome
                    continue
                if categoria_atual is None:
                    continue
                conta_b = ws.Cells(r, 2).Value
                tipo_c = ws.Cells(r, 3).Value
                if conta_b is None or tipo_c not in ("V", "F"):
                    continue
                try:
                    conta_int = int(conta_b)
                except (TypeError, ValueError):
                    continue
                mapa[(conta_int, tipo_c)] = categoria_atual
            mapa_por_unidade[aba] = mapa
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()

    todas_chaves = set()
    for m in mapa_por_unidade.values():
        todas_chaves |= set(m.keys())
    mapa_referencia = {}
    for chave in todas_chaves:
        valores = {m.get(chave) for m in mapa_por_unidade.values() if m.get(chave) is not None}
        if len(valores) == 1:
            mapa_referencia[chave] = valores.pop()
        else:
            log(f"AVISO: {chave} tem categoria DIFERENTE entre unidades no arquivo antigo: {valores}")
    return mapa_referencia


def classificar_hoje(conta_int: int, tipo: str, conta_geral) -> str:
    """Mesma lógica de ler_e_classificar (gerar_rateio_custos.py) - repetida
    aqui (não importada direto) pra essa ferramenta de auditoria não
    depender de reestruturar o Passo 5 só pra expor a decisão por linha."""
    if conta_int in CONTAS_FORCADAS_DEPRECIATION:
        return "Depreciation"
    if tipo == "V" and conta_int in CONTAS_FORCADAS_HANDLING_VARIAVEL:
        return "Handling"
    if tipo == "V" and conta_int in CONTAS_FORCADAS_TRANSPORTATION_VARIAVEL:
        return "Transportation"
    subcat = _resolver_subcategoria(tipo, conta_geral)
    if subcat is None:
        return "Other Variable" if tipo == "V" else "Other Fixed"
    return subcat


def coletar_ajs_vistos(meses: list[int], ano: int, ciclos: list[str], log=print) -> dict:
    """Devolve {(conta, tipo): {valores de AJ vistos}} varrendo a Base
    Intermediária de todos os mes/ciclo pedidos."""
    ajs_por_conta_tipo = defaultdict(set)
    for mes in meses:
        for ciclo in ciclos:
            try:
                caminho_bi = localizar_base_intermediaria(mes, ano, ciclo)
            except FileNotFoundError:
                log(f"AVISO: Base Intermediária de {mes:02d}/{ano} {ciclo} não encontrada - pulando.")
                continue
            wb = openpyxl.load_workbook(caminho_bi, read_only=True, data_only=True, keep_links=False)
            ws = wb["Intermediária"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=36):
                conta = row[0].value
                if conta is None:
                    continue
                try:
                    conta_int = int(conta)
                except (TypeError, ValueError):
                    continue
                tipo = row[26].value
                if tipo not in ("V", "F"):
                    continue
                ajs_por_conta_tipo[(conta_int, tipo)].add(row[35].value)
    return ajs_por_conta_tipo


def checar_classificacao(
    caminho_abertura_custos: Path, meses: list[int], ano: int, ciclos: list[str] | None = None, log=print
) -> dict:
    """Compara, conta por conta, a categoria declarada no arquivo antigo
    contra a que a classificação atual (AA/AJ + exceções) produziria hoje,
    testando TODOS os valores de AJ realmente vistos nos mes/ciclo pedidos.
    Devolve {"confere": [...], "diverge": [...], "sem_dado": [...]}."""
    ciclos = ciclos or ["Actual", "Flash"]
    mapa_referencia = extrair_estrutura_arquivo_antigo(caminho_abertura_custos, log)
    log(f"Estrutura extraída do arquivo antigo: {len(mapa_referencia)} combinações (conta, tipo).")

    ajs_por_conta_tipo = coletar_ajs_vistos(meses, ano, ciclos, log)

    confere, diverge, sem_dado = [], [], []
    for (conta, tipo), categoria_esperada in sorted(mapa_referencia.items()):
        ajs = ajs_por_conta_tipo.get((conta, tipo))
        if not ajs:
            sem_dado.append((conta, tipo, categoria_esperada))
            continue
        resultados = {classificar_hoje(conta, tipo, aj) for aj in ajs}
        if resultados == {categoria_esperada}:
            confere.append((conta, tipo, categoria_esperada))
        else:
            diverge.append((conta, tipo, categoria_esperada, resultados, ajs))

    log(f"BATE 100%: {len(confere)}")
    log(f"Sem dado suficiente pra testar (conta sem movimento nos mes/ciclo pedidos): {len(sem_dado)}")
    log(f"DIVERGE: {len(diverge)}")
    for d in diverge:
        log(f"  DIVERGE: conta={d[0]} tipo={d[1]} esperado={d[2]!r} classificação_hoje={d[3]!r} (AJ vistos: {d[4]!r})")

    return {"confere": confere, "diverge": diverge, "sem_dado": sem_dado}


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Audita a classificação Variável/Fixo do Passo 5 contra o arquivo antigo")
    parser.add_argument("--arquivo-antigo", type=Path, required=True, help="Caminho do '_Abertura custos...xlsx' de referência")
    parser.add_argument("--meses", type=int, nargs="+", required=True)
    parser.add_argument("--ano", type=int, required=True)
    parser.add_argument("--ciclos", nargs="+", default=["Actual", "Flash"])
    args = parser.parse_args()

    checar_classificacao(args.arquivo_antigo, args.meses, args.ano, args.ciclos)
