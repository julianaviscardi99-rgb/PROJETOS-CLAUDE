#!/usr/bin/env python3
"""
Confere se as contas contabeis do arquivo "Sem Agrupamento" da KSB1 (Fitted
Units) estao todas vinculadas a um agrupamento gestorial, comparando com o
arquivo "Gestoriais" do mesmo mes.

Regra de negocio (confirmada pela Juliana em 2026-08-10):
- Check 1: no "Sem Agrupamento", ignorar as contas contabeis que nao entram
  nessa comparacao porque nunca vao ter agrupamento gestorial (contas fixas
  da lista IGNORAR_EXATAS + qualquer conta que comece com "B").
- Check 2: das contas que sobraram, verificar se todas aparecem no arquivo
  "Gestoriais". As que nao aparecerem precisam ser mandadas para o time de
  controladoria central vincular ao agrupamento gestorial.
- Soma: o total do "Sem Agrupamento" (ja excluindo as contas do Check 1) deve
  bater com o total do "Gestoriais" se toda conta estiver vinculada.
- As linhas de subtotal do KSB1 (preenchimento amarelo, sem conta contabil)
  sao ignoradas nas somas.

Gera um arquivo "Check de agrupamentos - MM.YYYY.xlsx" na mesma pasta de rede
do mes, com o resultado dos dois checks e o resumo da soma.
"""
from pathlib import Path

from openpyxl import Workbook, load_workbook

from atualizar_ksb1_gui import BU, MESES_PASTA, REDE_BASE, nome_com_versao

IGNORAR_EXATAS = {
    "C23020JJ15",
    "K610100000",
    "M120400001",
    "M12040J001",
    "M130120110",
    "B220400000",
}

COL_CONTA = 4   # "Classe de custo"
COL_VALOR = 17  # "Valor/MR"

AMARELO_INDEXED = {13}
AMARELO_RGB = {"FFFFFF00", "00FFFF00", "FFFFFF00FF"}


def eh_conta_ignorada(conta: str) -> bool:
    return conta in IGNORAR_EXATAS or conta.upper().startswith("B")


def linha_e_subtotal(ws, r: int) -> bool:
    fill = ws.cell(row=r, column=1).fill
    if fill.patternType != "solid":
        return False
    fg = fill.fgColor
    if fg.type == "indexed" and fg.indexed in AMARELO_INDEXED:
        return True
    if fg.type == "rgb" and fg.rgb in AMARELO_RGB:
        return True
    return False


def ler_linhas_detalhe(caminho: Path):
    """Le (conta, valor) de cada linha de detalhe, pulando subtotais e linhas sem conta."""
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active
    linhas = []
    for r in range(2, ws.max_row + 1):
        conta = ws.cell(row=r, column=COL_CONTA).value
        if conta in (None, ""):
            continue
        if linha_e_subtotal(ws, r):
            continue
        valor = ws.cell(row=r, column=COL_VALOR).value or 0
        linhas.append((str(conta).strip(), float(valor)))
    return linhas


def encontrar_arquivo(pasta: Path, prefixo: str) -> Path:
    candidatos = sorted(pasta.glob(f"{prefixo}*.XLSX"), key=lambda p: p.stat().st_mtime)
    if not candidatos:
        raise FileNotFoundError(f"Não encontrei nenhum arquivo começando com '{prefixo}' em {pasta}")
    return candidatos[-1]


def gerar_check(mes: int, ano: int, log=print) -> Path:
    pasta_mes = REDE_BASE / str(ano) / "00.Extração Base KSB1" / MESES_PASTA[mes]

    arquivo_gest = encontrar_arquivo(pasta_mes, f"KSB1 - {BU['nome']} {mes:02d}.{ano} - Gestoriais")
    arquivo_sem = encontrar_arquivo(pasta_mes, f"KSB1 - {BU['nome']} {mes:02d}.{ano} - Sem Agrupamento")
    log(f"Comparando:\n  Gestoriais: {arquivo_gest.name}\n  Sem Agrupamento: {arquivo_sem.name}")

    linhas_sem = ler_linhas_detalhe(arquivo_sem)
    linhas_gest = ler_linhas_detalhe(arquivo_gest)
    contas_gest = {conta for conta, _ in linhas_gest}

    soma_por_conta_sem = {}
    linhas_filtradas = []
    for conta, valor in linhas_sem:
        soma_por_conta_sem[conta] = soma_por_conta_sem.get(conta, 0.0) + valor
        if not eh_conta_ignorada(conta):
            linhas_filtradas.append((conta, valor))

    contas_faltando = {}
    for conta, valor in linhas_filtradas:
        if conta not in contas_gest:
            contas_faltando[conta] = contas_faltando.get(conta, 0.0) + valor

    total_sem_filtrado = sum(v for _, v in linhas_filtradas)
    total_gest = sum(v for _, v in linhas_gest)
    diferenca = total_sem_filtrado - total_gest

    # Check 1 sempre lista as contas fixas pedidas pela usuaria, mesmo que
    # nao apareçam no mes (valor 0), mais qualquer conta comecando com "B"
    # que de fato apareceu no arquivo (essa parte da regra e um padrao, nao
    # uma lista fechada de contas).
    contas_b_encontradas = sorted(
        c for c in soma_por_conta_sem if c not in IGNORAR_EXATAS and c.upper().startswith("B")
    )

    wb_out = Workbook()
    ws = wb_out.active
    ws.title = f"Check {mes:02d}.{ano}"[:31]

    ws.append(["Check 1: contas ignoradas (nunca têm agrupamento gestorial)"])
    ws.append(["Conta contábil", "Valor no Sem Agrupamento"])
    for conta in sorted(IGNORAR_EXATAS):
        ws.append([conta, soma_por_conta_sem.get(conta, 0.0)])
    for conta in contas_b_encontradas:
        ws.append([conta, soma_por_conta_sem[conta]])

    ws.append([])
    if contas_faltando:
        ws.append(["Check 2: contas contábeis SEM vínculo no agrupamento gestorial — enviar para a controladoria central"])
        ws.append(["Conta contábil", "Valor total"])
        for conta, valor in sorted(contas_faltando.items()):
            ws.append([conta, valor])
    else:
        ws.append(["Check 2: todas as contas contábeis do Sem Agrupamento (após o Check 1) estão no agrupamento gestorial."])

    ws.append([])
    ws.append(["Resumo da conferência", f"{mes:02d}.{ano}"])
    ws.append(["Total Sem Agrupamento (excluindo contas do Check 1)", total_sem_filtrado])
    ws.append(["Total Gestoriais", total_gest])
    ws.append(["Diferença", diferenca])
    ws.append([
        "Situação",
        "OK - valores batem" if abs(diferenca) < 0.01 else "ATENÇÃO - valores não batem, ver Check 2",
    ])

    nome_saida = nome_com_versao(pasta_mes, f"Check de agrupamentos - {mes:02d}.{ano}.xlsx")
    caminho_saida = pasta_mes / nome_saida
    wb_out.save(caminho_saida)
    log(f"\nArquivo gerado: {caminho_saida}")
    log(f"Total Sem Agrupamento (filtrado): {total_sem_filtrado:,.2f}")
    log(f"Total Gestoriais: {total_gest:,.2f}")
    log(f"Diferença: {diferenca:,.2f}")
    if contas_faltando:
        log(f"AVISO: {len(contas_faltando)} conta(s) sem vínculo no gestorial (ver Check 2 no arquivo).")
    return caminho_saida


if __name__ == "__main__":
    import sys

    if len(sys.argv) != 3:
        print("Uso: python check_agrupamentos_ksb1.py <mes> <ano>")
        sys.exit(1)
    gerar_check(int(sys.argv[1]), int(sys.argv[2]))
