#!/usr/bin/env python3
"""
Analisa a ZLFIB (Pesquisa genérica de Notas Fiscais) em busca de Notas Fiscais
duplicadas — projeto "Fitted Recuperação".

Regras confirmadas pela Juliana em 2026-08-11:
- "Lista por Item" (não "Lista por Nota", que não funciona bem nessa transação).
  Cada NF aparece repetida uma vez por item — precisa agrupar por Nr Documento
  (DOCNUM) para voltar ao nível de NF antes de procurar duplicidade.
- Filiais da Fitted Units: 0031=SJP, 0032=IBI, 0053=SOR, 0054=GOI.
- "Buscar chave de acesso" (ACKEY) marcado: notas de mercadoria têm chave de
  acesso (identificador único da NF de verdade — se a mesma chave aparece em
  mais de um Nr Documento, é duplicidade certa). Notas de serviço NÃO têm
  chave de acesso — nesses casos, duplicidade é inferida por
  Parceiro + Nota Fiscal + Série + Valor total repetidos em mais de um
  Nr Documento.

Regras adicionadas em 2026-08-13:
- Direção do movimento (S_DIRECT) = '1' (Entrada): só interessam notas de
  entrada de fornecedor, não saídas. Reduz bastante o volume de linhas
  (testado em SJP/jan-2026: 381 -> 175 linhas de item).
- Filtro por "operação A24" (transferência de material) pedido pela Juliana
  NÃO foi localizado na tela da ZLFIB (não é Cfop nem Tipo NF — os dois
  campos foram checados ao vivo e não têm esse código). Decisão: seguir
  sem esse filtro por enquanto; ela ainda vai confirmar onde fica o campo
  "OPERA" que ela usa no dia a dia.
- Cruzamento com o estudo de duplicidade da KSB1 (projeto irmão, ver
  analisar_duplicidade_pagamento.py): os fornecedores que já apareceram
  como duplicados na KSB1 (por Documento de compras ou por Data) são
  marcados nas notas da ZLFIB, como sinal cruzado de confiança.
- CORREÇÃO IMPORTANTE (2026-08-13): a versão original lia a grid célula a
  célula via COM (`grid.GetCellValue`), que se mostrou pouco confiável em
  grades grandes (testado: SJP com 1.348 linhas — a leitura por COM achou
  só 38 números de documento únicos; exportando a mesma consulta pra
  arquivo, o número real era 565). Ver memory/errors/2026-08-13_zlfib_
  getcellvalue_dados_incorretos.md. Agora o script exporta a grade pra um
  arquivo temporário via menu nativo do SAP (Lista > Exportar > Planilha
  eletrônica — mesmo mecanismo já usado em atualizar_ksb1_gui.py pra KSB1)
  e lê o arquivo com openpyxl, em vez de usar GetCellValue.
- Tipo NF 'R8' = transferência de material (confirmado pela Juliana em
  2026-08-13, com exemplo real: duas notas da FIAT AUTOMOVEIS S/A em GOI,
  mesma chave de acesso, valores diferentes — não é duplicidade de
  pagamento a fornecedor, é movimento de transferência). Excluído da
  análise de duplicidade a partir de agora.
- SOR e GOI são plantas diferentes (Sorocaba e Goiana) — nunca tratar como
  uma coisa só na análise ou no texto de resultado, mesmo quando rodadas
  juntas no mesmo lote por conveniência (a coluna "Filial" já distingue).
"""
import tempfile
import time
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from atualizar_ksb1_gui import connect_session

FILIAIS = {"0031": "SJP", "0032": "IBI", "0053": "SOR", "0054": "GOI"}
DATA_DE = "01.01.2026"
DATA_ATE = "31.07.2026"
DIRECAO_ENTRADA = "1"
NFTYPE_EXCLUIDOS = {"R8"}  # transferencia de material, nao e' duplicidade de pagamento

PASTA_DESTINO = Path(
    r"\\FSS024-01BR.group.pirelli.com\GFU_DAC\Custos Fitted Units\Estudos\Estudo Duplicidade Pagamento"
)
ARQUIVO_DUP_KSB1 = PASTA_DESTINO / "Análise Duplicidade Pagamento.xlsx"

DESTAQUE_FORNECEDOR = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def carregar_fornecedores_duplicados_ksb1(caminho: Path = ARQUIVO_DUP_KSB1, log=print) -> set:
    """Lê o resultado já gerado do estudo de duplicidade da KSB1
    (analisar_duplicidade_pagamento.py) e devolve o conjunto de códigos de
    fornecedor que apareceram em algum grupo duplicado lá (por Documento ou
    por Data), pra cruzar com as notas da ZLFIB."""
    if not caminho.exists():
        log(f"Aviso: {caminho.name} não encontrado — seguindo sem cruzamento com a KSB1.")
        return set()
    wb = load_workbook(caminho, data_only=True, read_only=True)
    fornecedores = set()
    for nome_aba in ("Dup. por Documento", "Dup. por Data"):
        if nome_aba not in wb.sheetnames:
            continue
        ws = wb[nome_aba]
        for row in ws.iter_rows(min_row=2, values_only=True):
            fornecedor = row[0]
            if fornecedor:
                fornecedores.add(str(fornecedor).strip())
    log(f"{len(fornecedores)} fornecedor(es) distinto(s) já duplicado(s) no estudo da KSB1 (cruzamento).")
    return fornecedores


def abrir_zlfib(session, log):
    log("Abrindo a transação ZLFIB...")
    session.FindById("wnd[0]/tbar[0]/okcd").Text = "/nZLFIB"
    session.FindById("wnd[0]").SendVKey(0)
    time.sleep(1)
    if session.Info.Transaction != "ZLFIB":
        raise RuntimeError(f"Não consegui abrir a ZLFIB (tela atual: '{session.Info.Transaction}').")


def buscar_filial(session, filial: str, data_de: str, data_ate: str, log):
    wnd = session.FindById("wnd[0]")
    wnd.FindById("usr/ctxtP_BUKRS").Text = "0580"
    wnd.FindById("usr/ctxtS_BRANCH-LOW").Text = filial
    wnd.FindById("usr/ctxtS_PSTDAT-LOW").Text = data_de
    wnd.FindById("usr/ctxtS_PSTDAT-HIGH").Text = data_ate
    wnd.FindById("usr/ctxtS_DIRECT-LOW").Text = DIRECAO_ENTRADA
    wnd.FindById("usr/radP_ITEM").Select()
    wnd.FindById("usr/chkP_ACKEY").Selected = True
    log(f"Executando ZLFIB — Filial {filial} ({FILIAIS[filial]}), só Entradas, {data_de} a {data_ate}...")
    wnd.SendVKey(8)
    time.sleep(2)


# Colunas do arquivo exportado (Lista > Exportar > Planilha eletrônica),
# na ordem em que a ZLFIB as gera — checado ao vivo em 2026-08-13.
COLUNAS_EXPORT = [
    "BRANCH", "DOCNUM", "NFNUM", "SERIES", "NFTYPE", "PSTDAT", "DOCDAT",
    "PARID", "CGC", "NAME1", "BRGEW", "REGIOD", "NFTOT", "IPIBASE",
    "ICMSBASE", "ICMSVAL", "OBSERVAT", "ACKEY", "CRENAM", "NRO_NOTA",
]


def exportar_grid_para_arquivo(session, filial: str, pasta_tmp: Path, log) -> Path:
    """Exporta a grade de resultado pra um .xlsx via menu nativo do SAP
    (Lista > Exportar > Planilha eletrônica) — mais confiável que ler célula
    a célula via COM em grades grandes (ver nota no topo do arquivo)."""
    nome_arquivo = f"_tmp_zlfib_{filial}.xlsx"
    caminho = pasta_tmp / nome_arquivo
    if caminho.exists():
        caminho.unlink()

    session.FindById("wnd[0]/mbar/menu[0]/menu[3]/menu[1]").Select()
    time.sleep(1)
    wnd1 = session.FindById("wnd[1]")
    wnd1.FindById("usr/ctxtDY_PATH").Text = str(pasta_tmp)
    wnd1.FindById("usr/ctxtDY_FILENAME").Text = nome_arquivo
    wnd1.FindById("tbar[0]/btn[0]").Press()

    for _ in range(40):
        if caminho.exists():
            break
        time.sleep(0.5)
    if not caminho.exists():
        raise RuntimeError(f"Exportação da ZLFIB (filial {filial}) não gerou o arquivo esperado em {caminho}.")
    return caminho


def ler_arquivo_exportado(caminho: Path, log) -> list:
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb.active
    linhas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # cabeçalho
        linha = dict(zip(COLUNAS_EXPORT, row))
        linha["NFTOT"] = float(linha["NFTOT"] or 0)
        linha["ACKEY"] = str(linha["ACKEY"] or "").strip()
        for campo in ("BRANCH", "DOCNUM", "NFNUM", "SERIES", "NFTYPE", "PARID", "NAME1"):
            linha[campo] = str(linha[campo]).strip() if linha[campo] is not None else ""
        linhas.append(linha)
    wb.close()
    log(f"  {len(linhas)} linha(s) de item lida(s) do arquivo exportado.")
    return linhas


def coletar_todas_filiais(filiais: dict, data_de: str, data_ate: str, log=print):
    session = connect_session()
    todas_linhas = []
    with tempfile.TemporaryDirectory(prefix="zlfib_export_", ignore_cleanup_errors=True) as pasta_tmp_str:
        pasta_tmp = Path(pasta_tmp_str)
        for filial in filiais:
            abrir_zlfib(session, log)
            buscar_filial(session, filial, data_de, data_ate, log)
            caminho = exportar_grid_para_arquivo(session, filial, pasta_tmp, log)
            linhas = ler_arquivo_exportado(caminho, log)
            todas_linhas.extend(linhas)
    return todas_linhas


def colapsar_por_nf(linhas_item: list) -> list:
    """Uma NF vem repetida uma linha por item (mesmo DOCNUM); reduz a uma
    linha por Nr Documento, já que os campos de cabeçalho (NFNUM, SERIES,
    NFTOT, ACKEY etc.) se repetem idênticos em todas as linhas do item."""
    por_doc = {}
    qtd_itens = defaultdict(int)
    for l in linhas_item:
        qtd_itens[l["DOCNUM"]] += 1
        por_doc.setdefault(l["DOCNUM"], l)
    notas = []
    for docnum, cab in por_doc.items():
        nota = dict(cab)
        nota["QTD_ITENS"] = qtd_itens[docnum]
        nota["valor"] = cab["NFTOT"]
        notas.append(nota)
    return notas


TAMANHO_CHAVE_ACESSO_NFE = 44  # chave de acesso real da NFe tem sempre 44 digitos;
# valores mais curtos (ex: "0000000084") sao lixo/campo em branco preenchido com
# outro numero, nao uma chave de verdade — achado em 2026-08-13 (falso positivo
# em SOR/GOI: duas notas de fornecedores diferentes "batendo" nesse valor curto).


def encontrar_duplicidades(notas: list):
    com_chave = [n for n in notas if len(n["ACKEY"]) == TAMANHO_CHAVE_ACESSO_NFE]
    sem_chave = [n for n in notas if len(n["ACKEY"]) != TAMANHO_CHAVE_ACESSO_NFE]

    grupos_chave = defaultdict(list)
    for n in com_chave:
        grupos_chave[n["ACKEY"]].append(n)
    dup_chave = {k: v for k, v in grupos_chave.items() if len({n["DOCNUM"] for n in v}) > 1}

    grupos_sem_chave = defaultdict(list)
    for n in sem_chave:
        chave = (n["PARID"], n["NFNUM"], n["SERIES"], round(n["valor"], 2))
        grupos_sem_chave[chave].append(n)
    dup_sem_chave = {
        k: v for k, v in grupos_sem_chave.items() if len({n["DOCNUM"] for n in v}) > 1
    }

    return dup_chave, dup_sem_chave


def montar_planilha(wb, titulo, grupos, fornecedores_ksb1, log=print):
    ws = wb.create_sheet(titulo[:31])
    ws.append([
        "Filial", "Nr Documento", "Nota Fiscal", "Série", "Tipo NF", "Data lançamento",
        "Parceiro", "Nome", "Valor total NF", "Qtd. itens", "Chave de acesso",
        "Fornecedor também duplicado na KSB1",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for chave, grupo in sorted(grupos.items(), key=lambda kv: -kv[1][0]["valor"] * len(kv[1])):
        for n in grupo:
            tambem_ksb1 = "Sim" if n["PARID"] in fornecedores_ksb1 else ""
            ws.append([
                n["BRANCH"], n["DOCNUM"], n["NFNUM"], n["SERIES"], n["NFTYPE"], n["PSTDAT"],
                n["PARID"], n["NAME1"], n["valor"], n["QTD_ITENS"], n["ACKEY"], tambem_ksb1,
            ])
            if tambem_ksb1:
                ws.cell(row=ws.max_row, column=12).fill = DESTAQUE_FORNECEDOR
    return ws


def analisar(filiais: dict = None, data_de: str = None, data_ate: str = None, log=print) -> dict:
    filiais = filiais or FILIAIS
    data_de = data_de or DATA_DE
    data_ate = data_ate or DATA_ATE
    fornecedores_ksb1 = carregar_fornecedores_duplicados_ksb1(log=log)

    linhas_item = coletar_todas_filiais(filiais, data_de, data_ate, log)
    notas_brutas = colapsar_por_nf(linhas_item)
    notas = [n for n in notas_brutas if n["NFTYPE"] not in NFTYPE_EXCLUIDOS]
    excluidas_transferencia = len(notas_brutas) - len(notas)
    log(f"\n{len(linhas_item)} linha(s) de item no total -> {len(notas_brutas)} Nota(s) Fiscal(is) únicas "
        f"({excluidas_transferencia} excluída(s) por ser transferência de material, tipo(s) {sorted(NFTYPE_EXCLUIDOS)}) "
        f"-> {len(notas)} nota(s) analisada(s).")

    dup_chave, dup_sem_chave = encontrar_duplicidades(notas)

    total_dup_chave = sum(n["valor"] for g in dup_chave.values() for n in g)
    total_dup_sem_chave = sum(n["valor"] for g in dup_sem_chave.values() for n in g)

    fornecedores_dup_zlfib = {n["PARID"] for g in dup_chave.values() for n in g}
    fornecedores_dup_zlfib |= {n["PARID"] for g in dup_sem_chave.values() for n in g}
    cruzamento = fornecedores_dup_zlfib & fornecedores_ksb1

    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo["A1"] = "Análise de Notas Fiscais duplicadas (ZLFIB) — Fitted Units"
    ws_resumo["A1"].font = Font(bold=True, size=13)
    ws_resumo.append(["Período analisado", f"{data_de} a {data_ate}"])
    ws_resumo.append(["Filiais", ", ".join(f"{k} ({v})" for k, v in filiais.items())])
    ws_resumo.append(["Direção", "Só Entradas (fornecedor)"])
    ws_resumo.append([])
    ws_resumo.append(["Linhas de item lidas", len(linhas_item)])
    ws_resumo.append(["Notas Fiscais únicas (após agrupar por Nr Documento)", len(notas_brutas)])
    ws_resumo.append(["Excluídas por transferência de material (Tipo NF R8)", excluidas_transferencia])
    ws_resumo.append(["Notas efetivamente analisadas", len(notas)])
    ws_resumo.append([])
    ws_resumo.append(["Duplicidade por Chave de acesso (notas de mercadoria — mais confiável)"])
    ws_resumo.append(["Grupos duplicados", len(dup_chave)])
    ws_resumo.append(["Notas envolvidas", sum(len(g) for g in dup_chave.values())])
    ws_resumo.append(["Valor total envolvido", total_dup_chave])
    ws_resumo.append([])
    ws_resumo.append(["Duplicidade por Parceiro+NF+Série+Valor (notas de serviço, sem chave de acesso)"])
    ws_resumo.append(["Grupos duplicados", len(dup_sem_chave)])
    ws_resumo.append(["Notas envolvidas", sum(len(g) for g in dup_sem_chave.values())])
    ws_resumo.append(["Valor total envolvido", total_dup_sem_chave])
    ws_resumo.append([])
    ws_resumo.append(["Cruzamento com o estudo de duplicidade da KSB1 (analisar_duplicidade_pagamento.py)"])
    ws_resumo.append(["Fornecedores duplicados na KSB1 (Documento ou Data)", len(fornecedores_ksb1)])
    ws_resumo.append(["Desses, também com NF duplicada na ZLFIB", len(cruzamento)])

    montar_planilha(wb, "Dup. por Chave de Acesso", dup_chave, fornecedores_ksb1, log)
    montar_planilha(wb, "Dup. sem Chave (serviço)", dup_sem_chave, fornecedores_ksb1, log)

    nome_base = "Análise Duplicidade NF (ZLFIB).xlsx"
    candidato = PASTA_DESTINO / nome_base
    versao = 2
    while candidato.exists():
        candidato = PASTA_DESTINO / f"Análise Duplicidade NF (ZLFIB)_v{versao}.xlsx"
        versao += 1
    wb.save(candidato)
    log(f"\nArquivo gerado: {candidato}")

    return {
        "arquivo": candidato,
        "tem_duplicidade": bool(dup_chave or dup_sem_chave),
        "grupos_dup_chave": len(dup_chave),
        "grupos_dup_sem_chave": len(dup_sem_chave),
        "notas_envolvidas": sum(len(g) for g in dup_chave.values()) + sum(len(g) for g in dup_sem_chave.values()),
        "valor_total": total_dup_chave + total_dup_sem_chave,
    }


if __name__ == "__main__":
    analisar(filiais={"0031": "SJP"})
