#!/usr/bin/env python3
"""
Analisa o extrato da KSB1 (Fitted Units, Sem Agrupamento, periodo completo)
em busca de lancamentos/pagamentos a fornecedor em duplicidade.

Regra de negocio (confirmada pela Juliana em 2026-08-11, projeto "Fitted
Recuperacao"):
- Duplicidade = mesmo Fornecedor + mesmo Valor + mesmo "Documento de compras"
  OU mesmo Fornecedor + mesmo Valor + mesma "Data de lançamento", aparecendo
  em mais de uma linha.
- Documentos estornados nao contam: um estorno aparece como um segundo
  lancamento do mesmo fornecedor, com o mesmo valor em sinal contrario, que
  se cancela com o original — esses pares sao excluidos da analise antes de
  procurar duplicidade (senao um estorno seria confundido com duplicidade).
- Linhas de subtotal do KSB1 (preenchimento amarelo, sem "Classe de custo")
  sao ignoradas, mesma regra do check_agrupamentos_ksb1.py.
- So considera linhas com Fornecedor preenchido (linhas sem fornecedor, como
  rateios de RH, nao fazem parte deste estudo).

Gera "Análise Duplicidade Pagamento.xlsx" na mesma pasta do extrato, com
abas de resumo, os dois tipos de duplicidade e os estornos identificados
(para auditoria/transparência de como cada linha foi tratada).
"""
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from check_agrupamentos_ksb1 import linha_e_subtotal

COL_DATA_LANC = 1
COL_CENTRO = 3
COL_CONTA = 4
COL_FORNECEDOR = 6
COL_NOME1 = 7
COL_DOC_COMPRAS = 10
COL_VALOR = 17
COL_DOC_REFERENCIA = 19


def ler_linhas(caminho: Path, log=print):
    log(f"Lendo {caminho.name} (pode levar um tempo, arquivo grande)...")
    wb = load_workbook(caminho, data_only=True, read_only=True)
    ws = wb.active

    linhas = []
    r = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        r += 1
        conta = row[COL_CONTA - 1]
        if conta in (None, ""):
            continue
        fornecedor = row[COL_FORNECEDOR - 1]
        if fornecedor in (None, ""):
            continue
        valor = row[COL_VALOR - 1] or 0
        linhas.append({
            "linha_planilha": r + 1,
            "data_lancamento": row[COL_DATA_LANC - 1],
            "centro_custo": row[COL_CENTRO - 1],
            "conta": str(conta).strip(),
            "fornecedor": str(fornecedor).strip(),
            "nome1": row[COL_NOME1 - 1],
            "doc_compras": row[COL_DOC_COMPRAS - 1],
            "valor": float(valor),
            "doc_referencia": row[COL_DOC_REFERENCIA - 1],
        })
    log(f"{len(linhas)} linha(s) com fornecedor (de {r} linhas de detalhe lidas, ignorando as sem fornecedor).")
    return linhas


def excluir_subtotais(caminho: Path, linhas: list, log=print) -> list:
    # A leitura em ler_linhas usa read_only (rapido, mas sem acesso a
    # formatacao de celula), entao a checagem de subtotal (que depende do
    # preenchimento amarelo) e feita numa segunda passada, so nas linhas que
    # sobraram apos o filtro de fornecedor — bem mais rapido que checar a
    # formatacao das 175 mil linhas.
    if not linhas:
        return linhas
    log("Conferindo linhas de subtotal (2ª passada, só nas linhas com fornecedor)...")
    wb = load_workbook(caminho, data_only=True)
    ws = wb.active
    resultado = [l for l in linhas if not linha_e_subtotal(ws, l["linha_planilha"])]
    removidas = len(linhas) - len(resultado)
    if removidas:
        log(f"{removidas} linha(s) de subtotal removida(s).")
    return resultado


def identificar_estornos(linhas: list, log=print):
    """Casa pares (mesmo fornecedor, mesmo |valor|, sinal oposto, mesmo
    Documento de compras OU mesmo Doc.de referência quando presentes) e
    devolve (linhas_sem_estorno, pares_estornados)."""
    grupos = defaultdict(list)
    for l in linhas:
        chave_doc = l["doc_compras"] or l["doc_referencia"] or ""
        chave = (l["fornecedor"], round(abs(l["valor"]), 2), chave_doc)
        grupos[chave].append(l)

    excluidos_ids = set()
    pares = []
    for chave, grupo in grupos.items():
        if chave[2] == "":
            continue  # sem documento pra vincular, nao arrisca casar por coincidencia de valor
        positivos = [l for l in grupo if l["valor"] > 0]
        negativos = [l for l in grupo if l["valor"] < 0]
        for pos, neg in zip(positivos, negativos):
            excluidos_ids.add(id(pos))
            excluidos_ids.add(id(neg))
            pares.append((pos, neg))

    restantes = [l for l in linhas if id(l) not in excluidos_ids]
    log(f"{len(pares)} par(es) de estorno identificado(s) e excluído(s) da análise ({len(pares) * 2} linhas).")
    return restantes, pares


def encontrar_duplicidades(linhas: list, chave_extra: str):
    grupos = defaultdict(list)
    for l in linhas:
        chave = (l["fornecedor"], round(l["valor"], 2), l[chave_extra])
        if l[chave_extra] in (None, ""):
            continue
        grupos[chave].append(l)
    return {chave: grupo for chave, grupo in grupos.items() if len(grupo) > 1}


def montar_planilha_duplicidade(wb, titulo, duplicidades, rotulo_chave):
    ws = wb.create_sheet(titulo[:31])
    ws.append([
        "Fornecedor", "Nome", "Valor", rotulo_chave, "Qtd. ocorrências",
        "Data de lançamento", "Centro de custo", "Conta", "Doc. de compras", "Doc. de referência",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for chave, grupo in sorted(duplicidades.items(), key=lambda kv: -abs(kv[0][1]) * len(kv[1])):
        for l in grupo:
            ws.append([
                l["fornecedor"], l["nome1"], l["valor"], chave[2], len(grupo),
                l["data_lancamento"], l["centro_custo"], l["conta"], l["doc_compras"], l["doc_referencia"],
            ])
    return ws


def montar_planilha_estornos(wb, pares):
    ws = wb.create_sheet("Estornos identificados")
    ws.append([
        "Fornecedor", "Nome", "Valor original", "Valor estorno",
        "Doc. de compras/referência", "Data original", "Data estorno",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for pos, neg in pares:
        ws.append([
            pos["fornecedor"], pos["nome1"], pos["valor"], neg["valor"],
            pos["doc_compras"] or pos["doc_referencia"],
            pos["data_lancamento"], neg["data_lancamento"],
        ])
    return ws


def analisar(caminho_extrato: Path, log=print) -> Path:
    linhas = ler_linhas(caminho_extrato, log)
    linhas = excluir_subtotais(caminho_extrato, linhas, log)
    linhas, pares_estorno = identificar_estornos(linhas, log)

    dup_doc = encontrar_duplicidades(linhas, "doc_compras")
    dup_data = encontrar_duplicidades(linhas, "data_lancamento")

    total_dup_doc = sum(l["valor"] for grupo in dup_doc.values() for l in grupo)
    total_dup_data = sum(l["valor"] for grupo in dup_data.values() for l in grupo)

    wb = Workbook()
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    ws_resumo.append(["Análise de duplicidade de pagamento a fornecedor — Fitted Units"])
    ws_resumo.append(["Período analisado", "01.01.2026 a 31.07.2026"])
    ws_resumo.append(["Arquivo de origem", caminho_extrato.name])
    ws_resumo.append([])
    ws_resumo.append(["Linhas com fornecedor analisadas", len(linhas) + len(pares_estorno) * 2])
    ws_resumo.append(["Pares de estorno excluídos (não contam como duplicidade)", len(pares_estorno)])
    ws_resumo.append([])
    ws_resumo.append(["Critério 1: Fornecedor + Valor + Documento de compras"])
    ws_resumo.append(["Grupos com duplicidade encontrados", len(dup_doc)])
    ws_resumo.append(["Linhas envolvidas", sum(len(g) for g in dup_doc.values())])
    ws_resumo.append(["Valor total envolvido", total_dup_doc])
    ws_resumo.append([])
    ws_resumo.append(["Critério 2: Fornecedor + Valor + Data de lançamento"])
    ws_resumo.append(["Grupos com duplicidade encontrados", len(dup_data)])
    ws_resumo.append(["Linhas envolvidas", sum(len(g) for g in dup_data.values())])
    ws_resumo.append(["Valor total envolvido", total_dup_data])
    ws_resumo["A1"].font = Font(bold=True, size=13)

    montar_planilha_duplicidade(wb, "Dup. por Documento", dup_doc, "Doc. de compras")
    montar_planilha_duplicidade(wb, "Dup. por Data", dup_data, "Data de lançamento")
    montar_planilha_estornos(wb, pares_estorno)

    pasta = caminho_extrato.parent
    nome_base = "Análise Duplicidade Pagamento.xlsx"
    candidato = pasta / nome_base
    versao = 2
    while candidato.exists():
        candidato = pasta / f"Análise Duplicidade Pagamento_v{versao}.xlsx"
        versao += 1

    wb.save(candidato)
    log(f"\nArquivo gerado: {candidato}")
    return candidato


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Uso: python analisar_duplicidade_pagamento.py <caminho do extrato KSB1>")
        sys.exit(1)
    analisar(Path(sys.argv[1]))
